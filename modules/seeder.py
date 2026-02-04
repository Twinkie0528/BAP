"""
Reference Data Seeder for Central Planning Platform (CPP)
=========================================================

This module seeds the Reference Tables from the Master Excel file:
- BudgetCodeRef: Valid budget codes from "GENERAL" sheet
- ChannelCategory: Channel categories from "BUDGET LIST" sheet
- ChannelActivity: Activities per channel from "BUDGET LIST" sheet
- CampaignType: Campaign types from "DATA VALIDATION" sheet
- ProductService: Products & Services from "DATA VALIDATION" sheet
- Approver: Authorized approvers from "DATA VALIDATION" sheet

Usage:
    from modules.seeder import seed_all_reference_data
    seed_all_reference_data("path/to/MD UNIT BUDGET FORMAT 2025.xlsx")

Author: CPP Development Team
"""

import re
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime

import openpyxl
from sqlmodel import select

from database import (
    get_session, 
    BudgetCodeRef, 
    ChannelCategory, 
    ChannelActivity,
    CampaignType,
    ProductService,
    Approver
)

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# =============================================================================
# KNOWN CATEGORIES (Hardcoded based on actual file structure)
# =============================================================================

# These are the main channel categories in the Master Excel file
# Format: (display_order, name_in_mongolian, english_alias)
KNOWN_CATEGORIES = [
    (1, "ТВ СУВАГ", "TV"),
    (2, "СОШИАЛ", "SOCIAL"),
    (3, "ВЭБСАЙТ", "WEBSITE"),
    (4, "ГАДНАХ СУРТАЛЧИЛГАА", "OOH"),
    (5, "ДОТОР СУРТАЛЧИЛГАА", "INDOOR"),
    (6, "СОНИН СЭТГҮҮЛ", "PRINT"),
    (7, "FM CУВАГ", "FM"),
    (8, "КИНО ТЕАТР", "CINEMA"),
    (9, "КОНТЕНТ ХИЙЦЛЭЛ", "CONTENT"),
    (10, "ХЭВЛЭМЭЛ СУРТАЛЧИЛГАА", "PRINTED_ADS"),
    (11, "ДОТООД СУРТАЛЧИЛГАА", "INTERNAL"),
]

# =============================================================================
# KNOWN ACTIVITIES PER CATEGORY (Hardcoded from Master Excel)
# =============================================================================

KNOWN_ACTIVITIES = {
    "ТВ СУВАГ": [
        # 1.1 Нэвтрүүлэг
        "Аман сурталчилгаа", "Мэдээ", "Зар", "Ярилцлага",
        # 1.2.2 Төрөл
        "Реклам", "Нэвтрүүлэг",
        # 1.2.2 ЦАЦАЛТ (TV Stations)
        "TV мониторинг", "MNB", "EDU", "Монгол HD", "Movie box", "Asian box",
        "Bloomberg", "Central TV", "UBS", "C1", "MN25", "TV5", "TV9", "SBN",
        "NTV", "TV8", "PSN", "ETV", "Малчин ТВ", "Channel 11", "Eagle", "Эко",
        "Seven", "Соён гэгээрүүлэгч", "Like HD", "Dream TV", "Орон нутаг"
    ],
    
    "СОШИАЛ": [
        # 2.1 НӨЛӨӨЛӨГЧ
        "Сошиал", "ПиАр",
        # 2.2 ИДЭВХЖҮҮЛЭЛТ
        "Хэрэглэгчийн идэвхжүүлэлт", "Харилцагчийн идэвхжүүлэлт",
        "Дижитал эрхийн бичиг", "Бүтээгдэхүүн үйлчилгээний эрх", "Биет бүтээгдэхүүн",
        # 2.1 ЦАЦАЛТ (Social Platforms)
        "Facebook boost", "Google boost", "Youtube boost", "Instagram boost",
        "Twitter boost", "LinkedIn boost", "Viber boost", "Тик Ток boost",
        "Facebook пейж", "Facebook групп", "Spotify boost"
    ],
    
    "ВЭБСАЙТ": [
        # 3.1 БЭЛТГЭЛ
        "Нийтлэл бичүүлэх", "Нийтлэл бичих", "Видео сурвалжлага", "Фото баннер",
        "Эх бэлтгэх", "Видео баннер", "Урамшууллын мэдээ", "Онцлох мэдээ оруулах",
        "Шинэ төрлийн сурталчилгаа", "Инфографик мэдээ", "Бүүст", "Quiz",
        # 3.2 ЦАЦАЛТ (Websites)
        "Gogo.mn", "Ikon.mn", "Caak.mn", "Ub.life", "Unread.today", "barilga.mn",
        "profile.mn", "News.mn", "Goolingoo.mn", "Bolod.mn", "mass.mn", "medee.mn",
        "shuud.mn", "Shuurhai.mn", "Шар сайт", "Сайтуудын холбоо", "zaluu.com",
        "unegui.mn", "xopom.com", "bolor toli.com"
    ],
    
    "ГАДНАХ СУРТАЛЧИЛГАА": [
        # 4.1 БЭЛТГЭЛ
        "Гудамжны самбар", "Автобусны буудал", "Тугт самбар", "Лед дэлгэц",
        "Өлгөлт, буулгалт", "Хэвлэл", "Түрээс"
    ],
    
    "ДОТОР СУРТАЛЧИЛГАА": [
        # 5.1 БЭЛТГЭЛ
        "Орцны самбар", "Лифтний лед", "Худалдааны төв", "Оффис доторх лед",
        "Convenience store", "Coffee shop", "PC", "Баннер байршуулах", "Видео реклам"
    ],
    
    "СОНИН СЭТГҮҮЛ": [
        # 6.1 ЦАЦАЛТ
        "Фото баннер", "Мэдэгдэл гаргах", "Нийтлэл гаргах"
    ],
    
    "FM CУВАГ": [
        # 7.1 ЦАЦАЛТ
        "Нэвтрүүлэг", "Задгай цацалт", "Агуулга өгөх", "Багц авах",
        "Аман сурталчилгаа", "Ярилцлага"
    ],
    
    "КИНО ТЕАТР": [
        # 7.1 ЦАЦАЛТ
        "Дотор сурталчилгаа", "Гадна сурталчилгаа", "Үзвэрийн өмнөх цацалт",
        "Танхимын лед цацалт", "Кассын дэлгэц", "Пос баннер",
        "Хэвлэмэл сурталчилгаа", "Баннер байршуулах", "Талбай түрээс"
    ],
    
    "КОНТЕНТ ХИЙЦЛЭЛ": [
        # 7.1 БЭЛТГЭЛ
        "Брэндинг TVC", "Урамшууллын TVC", "Платформын контент",
        "Educational content", "Reel & Story", "Live", "Зохиомжит нэвтрүүлэг",
        "Баримтад кино", "Music video"
    ],
    
    "ХЭВЛЭМЭЛ СУРТАЛЧИЛГАА": [
        # 10.1 БЭЛТГЭЛ
        "Офсет хэвлэл", "Тараах материал", "Стэнд", "Боршур", "Flyer",
        "Эрхийн бичиг", "Фото хэвлэл", "UV биет материал", "Слайд", "Стикер",
        "Даавуу", "Хулдаас", "Лазер", "PVC", "Акрил", "Бусад",
        "Бүүт хийцлэл", "Хөөсөнцөр", "Даавуун хэвлэл", "Хулдаасан хэвлэл", "3D хэвлэл"
    ],
    
    "ДОТООД СУРТАЛЧИЛГАА": [
        # 11.1 ЦАЦАЛТ
        "Юнивишн", "Ддэш", "BC", "IVR", "Voice SMS",
        "Логин скрийн", "Pop up баннер", "Welcome screen", "Channel banner", "Volume banner"
    ],
}

# =============================================================================
# CAMPAIGN TYPES (From DATA VALIDATION - TYPE column)
# =============================================================================

KNOWN_CAMPAIGN_TYPES = [
    "BPR | OPTIMIZATION & DEVELOPMENT",
    "BRAND STRATEGY",
    "TALENT DEVELOPMENT",
    "BRANDING CAMPAIGN",
    "EDUCATIONAL CAMPAIGN",
    "PLATFORM DEVELOPMENT",
    "SPONSORSHIP",
    "PARTNERSHIP",
    "ESG",
    "CORPORATE SOCIAL RESPONSIBILITY",
    "PUBLIC RELATION",
    "RESEARCH",
    "INNOVATION",
    "CUSTOMER & TRADE PROMOTION",
    "PRODUCT LAUNCH",
    "ACQ PROMOTION",
    "INTERNAL",
    "RETENTION",
    "UPSELL",
]

# =============================================================================
# PRODUCTS & SERVICES (From DATA VALIDATION - PRODUCT & SERVICE column)
# =============================================================================

KNOWN_PRODUCTS_SERVICES = [
    # Main Products
    ("GROUP", None),
    ("UNITEL", None),
    ("UNIVISION", None),
    ("LOOKTV", None),
    ("GER INTERNET", None),
    ("B2B", None),
    ("TOKI", None),
    ("UNITEL HUB", None),
    ("ХӨВ", None),
    ("E-WASTE", None),
    ("LTE/DATA", None),
    ("CARD", None),
    ("DEVICE", None),
    ("PPS", "INDIVIDUAL PRE-PAID PLAN"),
    ("PS", "INDIVIDUAL POST-PAID PLAN"),
    ("PREFIX", None),
    ("COVERAGE", None),
    ("KIDS", None),
    ("T FAMILY", None),
    ("INTERNATIONAL", "ROAMING & TOUR SIM"),
    ("PREMIUM SERVICE", None),
    ("DIGITAL CHANNEL", "APP & CHATBOT"),
    ("SALES CHANNEL", "BRANCH"),
    ("VAS", None),
    ("TRIPLE PLAY", None),
    ("IPTV VOD", None),
    ("IPTV VAS", None),
    ("IPTV DEVICE", None),
    ("IP77", None),
    ("U BRAND", None),
    ("DIGITAL PRODUCT", None),
    ("PS&PPS", None),
]

# =============================================================================
# APPROVERS (From DATA VALIDATION - APPROVE SIGNED column)
# =============================================================================

KNOWN_APPROVERS = [
    ("А.ЭНХЧИМЭГ", "МАРКЕТИНГ ХАРИУЦСАН ДЭД ЗАХИРАЛ", 1),
    ("Д.ЖАМЪЯНШАРАВ", "ГҮЙЦЭТГЭХ ЗАХИРАЛ", 2),
    ("Э.ЭНХЦАЦРАЛ", "ДИЖИТАЛ МАРКЕТИНГИЙН МЕНЕЖЕР", 3),
]

# Regex pattern for budget codes (e.g., MD-BRANDING-MB1-110010001)
BUDGET_CODE_PATTERN = re.compile(r'^[A-Z]{2,}-[A-Z]+-[A-Z0-9]+-\d+$', re.IGNORECASE)

# Regex pattern for category headers (e.g., "1. ТВ СУВАГ", "10. ХЭВЛЭМЭЛ")
CATEGORY_HEADER_PATTERN = re.compile(r'^\d{1,2}\.\s*(.+)$')


# =============================================================================
# SEEDER FUNCTIONS
# =============================================================================

def seed_budget_codes(excel_path: str, sheet_name: str = "GENERAL") -> int:
    """
    Seed BudgetCodeRef table from the GENERAL sheet.
    
    Args:
        excel_path: Path to Master Excel file
        sheet_name: Name of the sheet containing budget codes
        
    Returns:
        Number of codes seeded
    """
    logger.info(f"📋 Seeding Budget Codes from '{sheet_name}' sheet...")
    
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"❌ Failed to open Excel file: {e}")
        return 0
    
    if sheet_name not in wb.sheetnames:
        logger.error(f"❌ Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        wb.close()
        return 0
    
    ws = wb[sheet_name]
    codes_found = []
    
    # Scan Column B (index 1) starting from row 10
    for row_idx in range(10, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=2).value  # Column B
        
        if cell_value and isinstance(cell_value, str):
            cell_value = cell_value.strip()
            
            # Check if it matches budget code pattern
            if BUDGET_CODE_PATTERN.match(cell_value):
                # Try to get description from adjacent column
                description = ws.cell(row=row_idx, column=3).value  # Column C
                if description:
                    description = str(description).strip()
                
                codes_found.append({
                    "code": cell_value,
                    "description": description,
                })
    
    wb.close()
    
    # Save to database
    count = 0
    with get_session() as session:
        for code_data in codes_found:
            # Check if already exists
            existing = session.exec(
                select(BudgetCodeRef).where(BudgetCodeRef.code == code_data["code"])
            ).first()
            
            if not existing:
                budget_code = BudgetCodeRef(
                    code=code_data["code"],
                    description=code_data.get("description"),
                    year=2025,
                    is_active=True
                )
                session.add(budget_code)
                count += 1
        
        session.commit()
    
    logger.info(f"✅ Seeded {count} new budget codes (found {len(codes_found)} total)")
    return count


def seed_channel_categories() -> Dict[str, int]:
    """
    Seed ChannelCategory table with known categories.
    
    Returns:
        Dictionary mapping category name to ID
    """
    logger.info("📋 Seeding Channel Categories...")
    
    category_map = {}
    
    with get_session() as session:
        for display_order, name, english_alias in KNOWN_CATEGORIES:
            # Check if already exists
            existing = session.exec(
                select(ChannelCategory).where(ChannelCategory.name == name)
            ).first()
            
            if existing:
                category_map[name] = existing.id
                # Also map by partial match
                category_map[name.upper()] = existing.id
            else:
                category = ChannelCategory(
                    name=name,
                    description=english_alias,
                    display_order=display_order,
                    is_active=True
                )
                session.add(category)
                session.commit()
                session.refresh(category)
                category_map[name] = category.id
                category_map[name.upper()] = category.id
        
        session.commit()
    
    logger.info(f"✅ Seeded {len(KNOWN_CATEGORIES)} channel categories")
    return category_map


def seed_channel_activities(category_map: Dict[str, int]) -> int:
    """
    Seed ChannelActivity table from hardcoded KNOWN_ACTIVITIES.
    
    Args:
        category_map: Dictionary mapping category names to IDs
        
    Returns:
        Number of activities seeded
    """
    logger.info("📋 Seeding Channel Activities from hardcoded data...")
    
    count = 0
    with get_session() as session:
        for category_name, activities in KNOWN_ACTIVITIES.items():
            # Find category ID
            category_id = category_map.get(category_name)
            if not category_id:
                logger.warning(f"⚠️ Category not found: {category_name}")
                continue
            
            for activity_name in activities:
                # Check if already exists
                existing = session.exec(
                    select(ChannelActivity).where(
                        ChannelActivity.category_id == category_id,
                        ChannelActivity.name == activity_name
                    )
                ).first()
                
                if not existing:
                    activity = ChannelActivity(
                        category_id=category_id,
                        name=activity_name,
                        is_active=True
                    )
                    session.add(activity)
                    count += 1
        
        session.commit()
    
    total_activities = sum(len(acts) for acts in KNOWN_ACTIVITIES.values())
    logger.info(f"✅ Seeded {count} new activities (total defined: {total_activities})")
    return count


def seed_campaign_types() -> int:
    """
    Seed CampaignType table with known campaign types.
    
    Returns:
        Number of types seeded
    """
    logger.info("📋 Seeding Campaign Types...")
    
    count = 0
    with get_session() as session:
        for order, name in enumerate(KNOWN_CAMPAIGN_TYPES, 1):
            existing = session.exec(
                select(CampaignType).where(CampaignType.name == name)
            ).first()
            
            if not existing:
                campaign_type = CampaignType(
                    name=name,
                    display_order=order,
                    is_active=True
                )
                session.add(campaign_type)
                count += 1
        
        session.commit()
    
    logger.info(f"✅ Seeded {count} new campaign types (total defined: {len(KNOWN_CAMPAIGN_TYPES)})")
    return count


def seed_products_services() -> int:
    """
    Seed ProductService table with known products and services.
    
    Returns:
        Number of products seeded
    """
    logger.info("📋 Seeding Products & Services...")
    
    count = 0
    with get_session() as session:
        for order, (name, description) in enumerate(KNOWN_PRODUCTS_SERVICES, 1):
            existing = session.exec(
                select(ProductService).where(ProductService.name == name)
            ).first()
            
            if not existing:
                product = ProductService(
                    name=name,
                    description=description,
                    display_order=order,
                    is_active=True
                )
                session.add(product)
                count += 1
        
        session.commit()
    
    logger.info(f"✅ Seeded {count} new products (total defined: {len(KNOWN_PRODUCTS_SERVICES)})")
    return count


def seed_approvers() -> int:
    """
    Seed Approver table with known approvers.
    
    Returns:
        Number of approvers seeded
    """
    logger.info("📋 Seeding Approvers...")
    
    count = 0
    with get_session() as session:
        for name, position, level in KNOWN_APPROVERS:
            existing = session.exec(
                select(Approver).where(Approver.name == name)
            ).first()
            
            if not existing:
                approver = Approver(
                    name=name,
                    position=position,
                    approval_level=level,
                    is_active=True
                )
                session.add(approver)
                count += 1
        
        session.commit()
    
    logger.info(f"✅ Seeded {count} new approvers (total defined: {len(KNOWN_APPROVERS)})")
    return count


def seed_all_reference_data(excel_path: str = None) -> Dict[str, int]:
    """
    Seed all reference tables.
    
    Categories and Activities are seeded from hardcoded data.
    Budget codes can optionally be seeded from Excel file.
    
    Args:
        excel_path: Optional path to Master Excel file (for budget codes only)
        
    Returns:
        Dictionary with counts of seeded items
    """
    logger.info("=" * 60)
    logger.info("🚀 Starting Reference Data Seeding")
    logger.info("=" * 60)
    
    results = {}
    
    # 1. Seed Budget Codes (from Excel if provided)
    if excel_path and Path(excel_path).exists():
        results["budget_codes"] = seed_budget_codes(excel_path, sheet_name="GENERAL")
    else:
        logger.info("📋 Skipping budget codes (no Excel file provided)")
        results["budget_codes"] = 0
    
    # 2. Seed Channel Categories (hardcoded)
    category_map = seed_channel_categories()
    results["categories"] = len(KNOWN_CATEGORIES)
    
    # 3. Seed Activities (hardcoded)
    results["activities"] = seed_channel_activities(category_map)
    
    # 4. Seed Campaign Types (hardcoded)
    results["campaign_types"] = seed_campaign_types()
    
    # 5. Seed Products & Services (hardcoded)
    results["products"] = seed_products_services()
    
    # 6. Seed Approvers (hardcoded)
    results["approvers"] = seed_approvers()
    
    logger.info("=" * 60)
    logger.info("✅ Reference Data Seeding Complete!")
    logger.info(f"   Budget Codes: {results['budget_codes']}")
    logger.info(f"   Categories: {results['categories']}")
    logger.info(f"   Activities: {results['activities']}")
    logger.info(f"   Campaign Types: {results['campaign_types']}")
    logger.info(f"   Products/Services: {results['products']}")
    logger.info(f"   Approvers: {results['approvers']}")
    logger.info("=" * 60)
    
    return results


def clear_reference_data() -> None:
    """
    Clear all reference data tables. Use with caution!
    """
    logger.warning("⚠️ Clearing all reference data...")
    
    with get_session() as session:
        # Delete in order due to foreign keys
        session.exec(select(ChannelActivity)).delete()
        session.exec(select(ChannelCategory)).delete()
        session.exec(select(BudgetCodeRef)).delete()
        session.exec(select(CampaignType)).delete()
        session.exec(select(ProductService)).delete()
        session.exec(select(Approver)).delete()
        session.commit()
    
    logger.info("✅ All reference data cleared")


def get_reference_data_stats() -> Dict[str, int]:
    """
    Get counts of reference data in the database.
    """
    with get_session() as session:
        budget_codes = len(session.exec(select(BudgetCodeRef)).all())
        categories = len(session.exec(select(ChannelCategory)).all())
        activities = len(session.exec(select(ChannelActivity)).all())
        campaign_types = len(session.exec(select(CampaignType)).all())
        products = len(session.exec(select(ProductService)).all())
        approvers = len(session.exec(select(Approver)).all())
    
    return {
        "budget_codes": budget_codes,
        "categories": categories,
        "activities": activities,
        "campaign_types": campaign_types,
        "products": products,
        "approvers": approvers
    }


# =============================================================================
# CLI INTERFACE
# =============================================================================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python -m modules.seeder <path_to_master_excel>")
        print("       python -m modules.seeder --stats")
        print("       python -m modules.seeder --clear")
        sys.exit(1)
    
    if sys.argv[1] == "--stats":
        stats = get_reference_data_stats()
        print(f"📊 Reference Data Stats:")
        print(f"   Budget Codes: {stats['budget_codes']}")
        print(f"   Categories: {stats['categories']}")
        print(f"   Activities: {stats['activities']}")
    
    elif sys.argv[1] == "--clear":
        confirm = input("⚠️ This will delete all reference data. Type 'yes' to confirm: ")
        if confirm.lower() == 'yes':
            clear_reference_data()
        else:
            print("❌ Cancelled")
    
    else:
        excel_path = sys.argv[1]
        seed_all_reference_data(excel_path)
