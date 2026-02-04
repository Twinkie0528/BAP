"""
Dashboard Page - CPP Analytics & File Management
=================================================

Central Planning Platform Dashboard with:
- CPP Analytics (Charts, KPIs, Trends)
- Budget file management
- PDF export capability

Author: CPP Development Team
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timezone, timedelta
import io
import os

# Page configuration
st.set_page_config(
    page_title="CPP Dashboard",
    page_icon="📊",
    layout="wide"
)

# Import our modules
from config import FileStatus
from database import get_session, BudgetFile, User
from modules.jwt_auth import get_current_user_from_token
from modules.file_storage import read_excel_file, get_excel_file_path
from modules.pdf_converter import get_pdf_as_bytes, convert_excel_sheet_to_pdf
from modules.analytics import (
    get_budget_summary,
    get_budget_by_company,
    get_budget_by_month,
    get_top_campaigns,
    get_budget_efficiency,
    get_status_distribution
)
from modules.report_generator import export_cpp_report
from sqlmodel import select

# Mongolia timezone (UTC+8)
MONGOLIA_TZ = timezone(timedelta(hours=8))


def format_datetime(dt):
    """Format datetime to Mongolia timezone."""
    if dt is None:
        return "N/A"
    return dt.strftime("%Y-%m-%d %H:%M")


def generate_excel_pdf(file_data: dict, excel_path: str = None) -> bytes:
    """
    Generate PDF from Excel using Native Excel conversion.
    This is the BEST quality - uses Microsoft Excel's own PDF export.
    """
    if not excel_path or not os.path.exists(excel_path):
        return None
    
    try:
        pdf_bytes = get_pdf_as_bytes(excel_path, None)
        
        if pdf_bytes:
            return pdf_bytes
        else:
            st.warning("PDF хөрвүүлэлт амжилтгүй. Microsoft Excel суусан эсэхийг шалгана уу.")
            return None
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        st.error(f"PDF үүсгэхэд алдаа: {e}")
        return None


# =============================================================================
# BULK EXCEL EXPORT FUNCTION
# =============================================================================

def generate_bulk_excel_export(all_files_data: list, all_channel_totals: dict, all_channel_details: dict, grand_totals: dict) -> bytes:
    """
    Generate a comprehensive Excel file with all budget data.
    
    Creates an Excel workbook with multiple sheets:
    1. Summary - All campaigns overview with key metrics
    2. Channels - Aggregated budget by channel type
    3. Details - All line items across all files
    4. Campaigns - Campaign headers and metadata
    
    Args:
        all_files_data: List of dicts with file, header, sections, totals
        all_channel_totals: Dict of channel name -> total budget
        all_channel_details: Dict of channel name -> list of file data
        grand_totals: Dict with total_budget and actual_budget
    
    Returns:
        Excel file as bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output = io.BytesIO()
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    
    number_format = '#,##0'
    currency_format = '₮#,##0'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =========================================================================
    # SHEET 1: Summary - All campaigns overview
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = "📊 ТӨСВИЙН НЭГТГЭЛ - БҮРЭН ТАЙЛАН"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    
    # Grand totals
    ws_summary['A3'] = "Нийт файлын тоо:"
    ws_summary['B3'] = len(all_files_data)
    ws_summary['C3'] = "Нийт төсөв:"
    ws_summary['D3'] = grand_totals.get('total_budget', 0)
    ws_summary['D3'].number_format = currency_format
    ws_summary['E3'] = "Нийт бодит:"
    ws_summary['F3'] = grand_totals.get('actual_budget', 0)
    ws_summary['F3'].number_format = currency_format
    
    for col in ['A3', 'C3', 'E3']:
        ws_summary[col].font = Font(bold=True)
    
    # Headers
    summary_headers = ["№", "Кампанит ажил", "Төсвийн код", "Компани", "Огноо", "Хугацаа", "Нийт төсөв", "Бодит төсөв"]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_idx, fd in enumerate(all_files_data, 6):
        header = fd['header']
        totals = fd['totals']
        f = fd['file']
        
        row_data = [
            row_idx - 5,
            header['campaign_name'] or "-",
            header['budget_code'] or f.budget_code or "-",
            header['company'] or "-",
            header['date'] or "-",
            header['period'] or "-",
            totals['total_budget'],
            totals['actual_budget']
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in [7, 8]:  # Budget columns
                cell.number_format = currency_format
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 5
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 25
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    ws_summary.column_dimensions['F'].width = 20
    ws_summary.column_dimensions['G'].width = 18
    ws_summary.column_dimensions['H'].width = 18
    
    # =========================================================================
    # SHEET 2: Channels - Budget by channel type
    # =========================================================================
    ws_channels = wb.create_sheet("Channels")
    
    ws_channels.merge_cells('A1:D1')
    ws_channels['A1'] = "📊 СУВГААР НЭГТГЭСЭН ТӨСӨВ"
    ws_channels['A1'].font = Font(bold=True, size=14)
    ws_channels['A1'].alignment = Alignment(horizontal="center")
    
    channel_headers = ["№", "Суваг", "Нийт төсөв", "Файлын тоо"]
    for col_idx, header in enumerate(channel_headers, 1):
        cell = ws_channels.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    sorted_channels = sorted(all_channel_totals.items(), key=lambda x: x[1], reverse=True)
    for row_idx, (ch_name, ch_total) in enumerate(sorted_channels, 4):
        num_files = len(all_channel_details.get(ch_name, []))
        row_data = [row_idx - 3, ch_name, ch_total, num_files]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_channels.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 3:
                cell.number_format = currency_format
    
    # Total row
    total_row = len(sorted_channels) + 4
    ws_channels.cell(row=total_row, column=1, value="").border = thin_border
    ws_channels.cell(row=total_row, column=2, value="НИЙТ").font = Font(bold=True)
    ws_channels.cell(row=total_row, column=2, value="НИЙТ").border = thin_border
    total_cell = ws_channels.cell(row=total_row, column=3, value=sum(all_channel_totals.values()))
    total_cell.number_format = currency_format
    total_cell.font = Font(bold=True)
    total_cell.border = thin_border
    ws_channels.cell(row=total_row, column=4, value=len(all_files_data)).border = thin_border
    
    ws_channels.column_dimensions['A'].width = 5
    ws_channels.column_dimensions['B'].width = 40
    ws_channels.column_dimensions['C'].width = 20
    ws_channels.column_dimensions['D'].width = 15
    
    # =========================================================================
    # SHEET 3: Details - All line items with full details
    # =========================================================================
    ws_details = wb.create_sheet("Details")
    
    ws_details.merge_cells('A1:L1')
    ws_details['A1'] = "📋 БҮРЭН ЗАДАРГАА - БҮХ МӨРҮҮД"
    ws_details['A1'].font = Font(bold=True, size=14)
    ws_details['A1'].alignment = Alignment(horizontal="center")
    
    detail_headers = [
        "№", "Кампанит ажил", "Суваг", "Дэд суваг", "Төрөл", "Хийгдэх ажил", 
        "Гүйцэтгэгч", "Хариуцагч", "Давтамж", "Нэгж үнэ", "Нийт төсөв", "Тайлбар"
    ]
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    detail_row = 4
    global_no = 1
    
    for fd in all_files_data:
        header = fd['header']
        sections = fd['sections']
        campaign_name = header['campaign_name'] or fd['file'].budget_code or fd['file'].filename
        
        for section in sections:
            section_name = section['name']
            
            # Process subsections
            if section['subsections']:
                for sub in section['subsections']:
                    subsection_name = sub['name']
                    for r in sub['rows']:
                        if not isinstance(r, dict):
                            continue
                        row_data = [
                            global_no,
                            campaign_name,
                            section_name,
                            subsection_name,
                            r.get('type', ''),
                            r.get('task', ''),
                            r.get('vendor', ''),
                            r.get('owner', ''),
                            r.get('freq', ''),
                            r.get('unit_price', ''),
                            r.get('total', 0) if r.get('total') else 0,
                            r.get('note', '')
                        ]
                        
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws_details.cell(row=detail_row, column=col_idx, value=value)
                            cell.border = thin_border
                            if col_idx == 11:  # Total column
                                cell.number_format = currency_format
                        
                        detail_row += 1
                        global_no += 1
            
            # Process direct rows (no subsection)
            elif section['rows']:
                for r in section['rows']:
                    if not isinstance(r, dict):
                        continue
                    row_data = [
                        global_no,
                        campaign_name,
                        section_name,
                        "-",
                        r.get('type', ''),
                        r.get('task', ''),
                        r.get('vendor', ''),
                        r.get('owner', ''),
                        r.get('freq', ''),
                        r.get('unit_price', ''),
                        r.get('total', 0) if r.get('total') else 0,
                        r.get('note', '')
                    ]
                    
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_details.cell(row=detail_row, column=col_idx, value=value)
                        cell.border = thin_border
                        if col_idx == 11:
                            cell.number_format = currency_format
                    
                    detail_row += 1
                    global_no += 1
    
    # Column widths for Details
    detail_widths = [5, 30, 25, 15, 12, 30, 20, 15, 10, 12, 15, 25]
    for idx, width in enumerate(detail_widths, 1):
        ws_details.column_dimensions[chr(64 + idx)].width = width
    
    # =========================================================================
    # SHEET 4: Campaigns - Full campaign metadata
    # =========================================================================
    ws_campaigns = wb.create_sheet("Campaigns")
    
    ws_campaigns.merge_cells('A1:J1')
    ws_campaigns['A1'] = "📝 КАМПАНИТ АЖЛЫН МЭДЭЭЛЭЛ"
    ws_campaigns['A1'].font = Font(bold=True, size=14)
    ws_campaigns['A1'].alignment = Alignment(horizontal="center")
    
    campaign_headers = [
        "№", "Кампанит ажил", "Төсвийн код", "Маркетингийн код", "Компани",
        "Зорилго", "Зорилтот хэрэглэгчид", "Голлох мессеж", "Хугацаа", "Батлагдсан"
    ]
    for col_idx, header in enumerate(campaign_headers, 1):
        cell = ws_campaigns.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, fd in enumerate(all_files_data, 4):
        header = fd['header']
        f = fd['file']
        
        row_data = [
            row_idx - 3,
            header['campaign_name'] or "-",
            header['budget_code'] or f.budget_code or "-",
            header['marketing_code'] or "-",
            header['company'] or "-",
            header['goal'] or "-",
            header['target_audience'] or "-",
            header['main_message'] or "-",
            header['period'] or "-",
            f"{header['approver']} {header['approver_name']}" if header['approver'] else "-"
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_campaigns.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    campaign_widths = [5, 35, 25, 20, 15, 35, 25, 35, 20, 25]
    for idx, width in enumerate(campaign_widths, 1):
        ws_campaigns.column_dimensions[chr(64 + idx)].width = width
    
    # =========================================================================
    # SHEET 5: Channel Details - Each channel with breakdown
    # =========================================================================
    ws_channel_details = wb.create_sheet("Channel Details")
    
    ws_channel_details.merge_cells('A1:G1')
    ws_channel_details['A1'] = "📊 СУВГИЙН ДЭЛГЭРЭНГҮЙ ЗАДАРГАА"
    ws_channel_details['A1'].font = Font(bold=True, size=14)
    ws_channel_details['A1'].alignment = Alignment(horizontal="center")
    
    ch_detail_headers = ["Суваг", "Кампанит ажил", "Дэд суваг", "Ажил", "Гүйцэтгэгч", "Нийт төсөв", "Тайлбар"]
    for col_idx, header in enumerate(ch_detail_headers, 1):
        cell = ws_channel_details.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ch_row = 4
    for ch_name in sorted(all_channel_totals.keys()):
        file_data_list = all_channel_details.get(ch_name, [])
        
        # Channel header row
        cell = ws_channel_details.cell(row=ch_row, column=1, value=ch_name)
        cell.font = subheader_font
        cell.fill = subheader_fill
        for col_idx in range(1, 8):
            ws_channel_details.cell(row=ch_row, column=col_idx).fill = subheader_fill
            ws_channel_details.cell(row=ch_row, column=col_idx).border = thin_border
        
        total_cell = ws_channel_details.cell(row=ch_row, column=6, value=all_channel_totals[ch_name])
        total_cell.number_format = currency_format
        total_cell.font = subheader_font
        ch_row += 1
        
        for file_data in file_data_list:
            section = file_data['section']
            campaign_name = file_data['file_name']
            
            if section['subsections']:
                for sub in section['subsections']:
                    for r in sub['rows']:
                        row_data = [
                            "",  # Channel already in header
                            campaign_name,
                            sub['name'],
                            r.get('task', ''),
                            r.get('vendor', ''),
                            r.get('total', 0) if r.get('total') else 0,
                            r.get('note', '')
                        ]
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws_channel_details.cell(row=ch_row, column=col_idx, value=value)
                            cell.border = thin_border
                            if col_idx == 6:
                                cell.number_format = currency_format
                        ch_row += 1
            
            elif section['rows']:
                for r in section['rows']:
                    row_data = [
                        "",
                        campaign_name,
                        "-",
                        r.get('task', ''),
                        r.get('vendor', ''),
                        r.get('total', 0) if r.get('total') else 0,
                        r.get('note', '')
                    ]
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_channel_details.cell(row=ch_row, column=col_idx, value=value)
                        cell.border = thin_border
                        if col_idx == 6:
                            cell.number_format = currency_format
                    ch_row += 1
    
    ch_detail_widths = [30, 35, 15, 35, 25, 18, 30]
    for idx, width in enumerate(ch_detail_widths, 1):
        ws_channel_details.column_dimensions[chr(64 + idx)].width = width
    
    # Save to BytesIO
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


# =============================================================================
# ANALYTICS TAB
# =============================================================================

def render_analytics_tab(session):
    """Render the CPP Analytics tab with charts and KPIs."""
    
    # Get summary data
    summary = get_budget_summary(session)
    
    if summary['file_count'] == 0:
        st.warning("📭 Одоогоор системд өгөгдөл байхгүй байна. Төсөв оруулна уу.")
        st.page_link("pages/2_📤_Upload.py", label="📤 Төсөв оруулах", icon="📤")
        return
    
    # KPI Row
    st.subheader("📊 Гол үзүүлэлтүүд")
    kpi1, kpi2, kpi3, kpi4 = st.columns(4)
    
    with kpi1:
        st.metric(
            "💰 Нийт төсөв", 
            f"₮{float(summary['total_planned']):,.0f}"
        )
    
    with kpi2:
        st.metric(
            "💵 Бодит төсөв", 
            f"₮{float(summary['total_actual']):,.0f}"
        )
    
    with kpi3:
        st.metric(
            "📁 Нийт файлууд", 
            f"{summary['file_count']} ш"
        )
    
    with kpi4:
        # Calculate budget difference (Planned - Actual)
        if summary['total_planned'] > 0:
            difference = float(summary['total_planned']) - float(summary['total_actual'])
            delta_color = "normal" if difference >= 0 else "inverse"
            st.metric(
                "📊 Төсвийн зөрүү",
                f"₮{abs(difference):,.0f}",
                delta=f"{'+' if difference >= 0 else ''}{difference:,.0f}",
                delta_color=delta_color,
                help="Нийт төсөв - Бодит төсөв. Эерөг: хэмнэсэн, Сөрөг: хэтэрсэн"
            )
        else:
            st.metric("📊 Төсвийн зөрүү", "N/A")
    
    st.divider()
    
    # Charts Row 1: Company breakdown and Top Campaigns
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🏢 Компаниар харах")
        df_company = get_budget_by_company(session)
        
        if not df_company.empty:
            fig_pie = px.pie(
                df_company,
                values='ActualAmount',
                names='Company',
                title='Бодит төсвийн хуваарилалт',
                hole=0.4,
                color_discrete_sequence=px.colors.qualitative.Set2
            )
            fig_pie.update_traces(textposition='inside', textinfo='percent+label')
            fig_pie.update_layout(showlegend=True, legend=dict(orientation="h", yanchor="bottom", y=-0.2))
            st.plotly_chart(fig_pie, width='stretch')
        else:
            st.info("Компанийн мэдээлэл олдсонгүй")
    
    with col2:
        st.subheader("🏆 Топ 10 кампанит ажил")
        df_top = get_top_campaigns(session, limit=10)
        
        if not df_top.empty:
            fig_bar = px.bar(
                df_top,
                x='ActualAmount',
                y='Campaign',
                orientation='h',
                color='Company',
                title='Хамгийн их төсөвтэй ажлууд',
                text_auto='.2s',
                color_discrete_sequence=px.colors.qualitative.Pastel
            )
            fig_bar.update_layout(
                yaxis={'categoryorder': 'total ascending'},
                height=400
            )
            st.plotly_chart(fig_bar, width='stretch')
        else:
            st.info("Кампанит ажлын мэдээлэл олдсонгүй")
    
    st.divider()
    
    # Charts Row 2: Efficiency and Status
    col3, col4 = st.columns(2)
    
    with col3:
        st.subheader("📊 Төсвийн үр ашиг")
        df_efficiency = get_budget_efficiency(session)
        
        if not df_efficiency.empty:
            # Create gauge-like bar chart
            fig_eff = px.bar(
                df_efficiency,
                x='Efficiency',
                y='Campaign',
                orientation='h',
                color='Efficiency',
                color_continuous_scale=['green', 'yellow', 'orange', 'red'],
                range_color=[0, 150],
                title='Гүйцэтгэл vs Төлөвлөгөө (%)',
                text='Status'
            )
            fig_eff.add_vline(x=100, line_dash="dash", line_color="red", annotation_text="100%")
            fig_eff.update_layout(
                yaxis={'categoryorder': 'total ascending'},
                height=350,
                showlegend=False
            )
            st.plotly_chart(fig_eff, width='stretch')
        else:
            st.info("Үр ашгийн мэдээлэл олдсонгүй")
    
    with col4:
        st.subheader("📈 Статусаар харах")
        df_status = get_status_distribution(session)
        
        if not df_status.empty:
            status_colors = {
                'Хүлээгдэж байна': '#FFA500',
                'Батлагдсан': '#4CAF50',
                'PDF үүссэн': '#2196F3',
                'Гарын үсэг оруулсан': '#9C27B0',
                'Дууссан': '#607D8B',
                'Буцаагдсан': '#F44336'
            }
            
            fig_status = px.pie(
                df_status,
                values='Count',
                names='StatusName',
                title='Файлуудын статус'
            )
            fig_status.update_traces(textposition='inside', textinfo='value+percent')
            st.plotly_chart(fig_status, width='stretch')
        else:
            st.info("Статусын мэдээлэл олдсонгүй")
    
    # Detailed table
    st.divider()
    st.subheader("📋 Дэлгэрэнгүй задаргаа")
    
    df_company = get_budget_by_company(session)
    if not df_company.empty:
        # Format for display
        df_display = df_company.copy()
        df_display['PlannedAmount'] = df_display['PlannedAmount'].apply(lambda x: f"₮{x:,.0f}")
        df_display['ActualAmount'] = df_display['ActualAmount'].apply(lambda x: f"₮{x:,.0f}")
        df_display.columns = ['Компани', 'Нийт төсөв', 'Бодит', 'Файлын тоо']
        st.dataframe(df_display, width='stretch', hide_index=True)


# =============================================================================
# FILES TAB
# =============================================================================

def render_files_tab(session):
    """Render the Files list tab."""
    
    # Load all files with uploader data using JOIN
    statement = (
        select(BudgetFile, User)
        .outerjoin(User, BudgetFile.uploader_id == User.id)
        .order_by(BudgetFile.uploaded_at.desc())
    )
    results = session.exec(statement).all()
    
    if not results:
        st.info("Одоогоор ямар ч төсөв байхгүй байна.")
        st.page_link("pages/2_📤_Upload.py", label="📤 Төсөв оруулах", icon="📤")
        return
    
    # Extract files list
    files = [r[0] for r in results]
    uploader_map = {r[0].id: r[1] for r in results}
    
    # Status counts
    st.subheader("📈 Төлөвийн статистик")
    
    status_counts = {}
    for file in files:
        status = file.status.value if hasattr(file.status, 'value') else str(file.status)
        status_counts[status] = status_counts.get(status, 0) + 1
    
    num_cols = min(len(status_counts) + 1, 5)
    cols = st.columns(num_cols)
    
    with cols[0]:
        st.metric("Нийт файл", len(files))
    
    for i, (status, count) in enumerate(status_counts.items(), 1):
        if i >= num_cols:
            break
        status_labels = {
            "pending_approval": "Хүлээгдэж буй",
            "approved_for_print": "Батлагдсан",
            "rejected": "Буцаагдсан",
            "finalized": "Дууссан"
        }
        with cols[i]:
            st.metric(status_labels.get(status, status), count)
    
    st.divider()
    
    # Files table
    st.subheader("📋 Бүх файлууд")
    
    data = []
    for file in files:
        status = file.status.value if hasattr(file.status, 'value') else str(file.status)
        status_labels = {
            "pending_approval": "🕐 Хүлээгдэж буй",
            "approved_for_print": "✅ Батлагдсан",
            "rejected": "❌ Буцаагдсан",
            "finalized": "🏁 Дууссан"
        }
        
        budget_type_label = "Үндсэн" if (hasattr(file.budget_type, 'value') and file.budget_type.value == "primary") else "Нэмэлт"
        
        uploader_user = uploader_map.get(file.id)
        uploader = uploader_user.full_name if uploader_user else "Unknown"
        
        actual_str = f"₮{float(file.total_amount):,.0f}" if file.total_amount else "N/A"
        planned_str = f"₮{float(file.planned_amount):,.0f}" if file.planned_amount else "N/A"
        upload_date = format_datetime(file.uploaded_at)
        
        budget_code = getattr(file, 'budget_code', None) or f"#{file.id}"
        brand = getattr(file, 'brand', None) or "-"
        
        data.append({
            "Төсвийн код": budget_code,
            "Брэнд": brand,
            "Файлын нэр": file.filename,
            "Төрөл": budget_type_label,
            "Төлөв": status_labels.get(status, status),
            "Нийт төсөв": planned_str,
            "Бодит": actual_str,
            "Төсөв оруулсан": uploader,
            "Огноо": upload_date
        })
    
    df = pd.DataFrame(data)
    st.dataframe(df, height=400, width='stretch')


# =============================================================================
# BUDGET REPORT TAB - Excel-ээс ирсэн төсвийн дэлгэрэнгүй
# =============================================================================

def parse_budget_header(df: pd.DataFrame) -> dict:
    """Extract header information from Excel budget file."""
    header = {
        "approver": "",
        "approver_name": "",
        "approver_title": "",
        "campaign_name": "",
        "date": "",
        "budget_code": "",
        "marketing_code": "",
        "company": "",
        "goal": "",
        "target_audience": "",
        "main_message": "",
        "target": "",
        "period": "",
        "content_production": "",
        "channel_distribution": "",
        "event_level": "",
        "scope": "",
        "budget_ref": ""
    }
    
    for i in range(min(20, len(df))):
        row = df.iloc[i]
        row_vals = [str(v).strip() if pd.notna(v) else "" for v in row.values]
        row_text = ' '.join(row_vals)
        
        # R1: БАТЛАВ, approver name, title
        if 'БАТЛАВ' in row_text:
            header["approver"] = "БАТЛАВ"
            header["approver_name"] = row_vals[5] if len(row_vals) > 5 else ""
            header["approver_title"] = row_vals[8] if len(row_vals) > 8 else ""
        
        # R4: Campaign name and date - find column dynamically
        if 'ТӨСӨВ' in row_text and 'Огноо' in row_text:
            header["campaign_name"] = row_vals[1] if len(row_vals) > 1 else ""
            # Find date after 'Огноо:' label
            for j, v in enumerate(row_vals):
                if 'Огноо' in v and j + 1 < len(row_vals):
                    header["date"] = row_vals[j + 1]
                    break
        
        # Budget code - find value after label dynamically
        if 'Төсвийн код' in row_text:
            for j, v in enumerate(row_vals):
                if 'Төсвийн код' in v and j + 1 < len(row_vals):
                    header["budget_code"] = row_vals[j + 1]
                    break
        
        # Marketing code - find value after label dynamically
        if 'Маркетингийн код' in row_text:
            for j, v in enumerate(row_vals):
                if 'Маркетингийн код' in v and j + 1 < len(row_vals):
                    header["marketing_code"] = row_vals[j + 1]
                    break
            # Mark this row index for company detection in next row
            header["_marketing_code_row"] = i
        
        # Company detection moved to after loop - uses budget_code first letter
        
        # Goal
        if 'Зорилго:' in row_text or row_text.startswith('Зорилго'):
            header["goal"] = row_vals[1] if len(row_vals) > 1 else ""
        
        # Target audience
        if 'Зорилтот хэрэглэгчид' in row_text:
            header["target_audience"] = row_vals[1] if len(row_vals) > 1 else ""
        
        # Main message
        if 'Голлох мессеж' in row_text:
            header["main_message"] = row_vals[1] if len(row_vals) > 1 else ""
        
        # Target
        if 'Таргет:' in row_text:
            header["target"] = row_vals[1] if len(row_vals) > 1 else ""
        
        # Period
        if 'хугацаа:' in row_text.lower():
            header["period"] = row_vals[1] if len(row_vals) > 1 else ""
        
        # Content production / Channel distribution
        if 'Content production' in row_text:
            try:
                val = float(row_vals[3]) if len(row_vals) > 3 and row_vals[3] else 0
                header["content_production"] = f"{val*100:.0f}%"
            except:
                pass
        
        if 'Channel distribution' in row_text:
            try:
                val = float(row_vals[3]) if len(row_vals) > 3 and row_vals[3] else 0
                header["channel_distribution"] = f"{val*100:.0f}%"
            except:
                pass
        
        # Event level, scope, budget ref
        if 'Идэвхжүүлэлтийн түвшин' in row_text:
            header["event_level"] = row_vals[5] if len(row_vals) > 5 else ""
            header["scope"] = row_vals[7] if len(row_vals) > 7 else ""
            header["budget_ref"] = row_vals[9] if len(row_vals) > 9 else ""
    
    # Company detection from budget_code first letter (primary method)
    # A=Юнител, B=Юнивишн, G=Green Future, J=IVLBS, T=MPSC
    if not header["company"] and header.get("budget_code"):
        company_map = {
            'A': 'Юнител',
            'B': 'Юнивишн',
            'G': 'Green Future',
            'J': 'IVLBS',
            'T': 'MPSC'
        }
        company_code = header["budget_code"][0].upper() if header["budget_code"] else ''
        header["company"] = company_map.get(company_code, '')
    
    return header


def parse_excel_sections(df: pd.DataFrame) -> tuple:
    """
    Parse Excel DataFrame to extract budget sections.
    Returns (sections list, totals dict, column_map).
    """
    sections = []
    current_section = None
    current_subsection = None
    totals = {
        "total_budget": 0,
        "internal_expense": 0,
        "long_term_expense": 0,
        "actual_budget": 0
    }
    
    # Main section keywords
    MAIN_SECTIONS = [
        'ДИЖИТАЛ СУРТАЛЧИЛГААНЫ СУВАГ',
        'ВЭБСАЙТ СУРТАЛЧИЛГААНЫ СУВАГ',
        'КОНТЕНТ ХИЙЦЛЭЛ',
        'ДОТООД СУРТАЛЧИЛГААНЫ СУВАГ',
        'УРТ ХУГАЦААНЫ ГЭРЭЭТ НӨЛӨӨЛӨГЧ',
        'УРТ ХУГАЦААНЫ ГЭРЭЭТ ТВ СУВАГ',
        'УРТ ХУГАЦААНЫ ГЭРЭЭТ ГАДНА СУРТАЛЧИЛГААНЫ СУВАГ',
        'ТВ СУРТАЛЧИЛГААНЫ СУВАГ',
        'ТВ СУВАГ',
        'СОШИАЛ',
        'FM СУРТАЛЧИЛГААНЫ СУВАГ',
        'FM СУВАГ',
        'КИНО ТЕАТР СУРТАЛЧИЛГААНЫ СУВАГ',
        'КИНО ТЕАТР',
        'ГАДНА СУРТАЛЧИЛГААНЫ СУВАГ',
        'ГАДНАХ СУРТАЛЧИЛГАА',
        'ДОТОР СУРТАЛЧИЛГААНЫ СУВАГ',
        'ХЭВЛЭМЭЛ СУРТАЛЧИЛГАА',
        'СОНИН СЭТГҮҮЛ СУРТАЛЧИЛГААНЫ СУВАГ',
        'СОНИН СЭТГҮҮЛ',
        'АНИМЭЙШН СУРТАЛЧИЛГААНЫ СУВАГ',
        'МОДО ФОНИЙ СУРТАЛЧИЛГААНЫ СУВАГ',
        'БУСАД СУРТАЛЧИЛГААНЫ СУВАГ',
        'ХӨЛТӨЛ СУРТАЛЧИЛГАА',
        'АР ГА ХЭМЖЭЭ',
        'СУДАЛГАА ШИНЖИЛГЭЭ',
        'ХАМТЫН АЖИЛЛАГАА'
    ]
    
    # Subsection keywords
    SUBSECTIONS = ['НӨЛӨӨЛӨГЧ', 'ЦАЦАЛТ', 'БЭЛТГЭЛ', 'ИДЭВХЖҮҮЛЭЛТ', 'НЭВТРҮҮЛЭГ']
    
    # First, find column mapping from header row
    column_map = {
        'no': 1,           # №
        'type': 2,         # Төрөл
        'task': 3,         # Хийгдэх ажил
        'vendor': 4,       # Гүйцэтгэгч/Нөлөөлөгч
        'date': None,      # Хугацаа (optional)
        'owner': 5,        # Хариуцах эзэн
        'freq': 6,         # Давтамж
        'unit_price': 7,   # Нэгж үнэ
        'total': 8,        # Нийт төсөв
        'note': 9          # Тайлбар
    }
    
    # Find header row and update column mapping
    for i in range(min(30, len(df))):
        row = df.iloc[i]
        row_vals = [str(v).strip().upper() if pd.notna(v) else "" for v in row.values]
        row_text = ' '.join(row_vals)
        
        if '№' in row_text and ('СУВАГ' in row_text or 'ИДЭВХЖҮҮЛЭЛТ' in row_text):
            # Found header row - map columns dynamically
            for j, val in enumerate(row_vals):
                if val == '№':
                    column_map['no'] = j
                elif 'ТАЙЛБАР' in val:
                    column_map['note'] = j
                elif 'НИЙТ ТӨСӨВ' in val:
                    column_map['total'] = j
                elif 'НЭГЖ ҮНЭ' in val:
                    column_map['unit_price'] = j
                elif 'ДАВТАМЖ' in val:
                    column_map['freq'] = j
                elif 'ХАРИУЦАХ' in val or 'ЭЗЭН' in val:
                    column_map['owner'] = j
                elif 'ХУГАЦАА' in val:
                    column_map['date'] = j
            break
    
    # Track last non-empty values for merged cell handling (forward fill)
    last_owner = ""
    last_date = ""
    
    for i, row in df.iterrows():
        row_vals = [str(v).strip() if pd.notna(v) else "" for v in row.values]
        col1 = row_vals[column_map['no']] if len(row_vals) > column_map['no'] else ""
        col1_upper = col1.upper()
        row_text = ' '.join(row_vals).upper()
        
        # Skip empty rows
        if not col1:
            continue
        
        # Skip header row
        if col1 == '№' or 'ИДЭВХЖҮҮЛЭЛТИЙН СУВАГ' in row_text:
            continue
        
        # IMPORTANT: Check section totals FIRST (before main section detection)
        # This prevents "НИЙТ КОНТЕНТ ХИЙЦЛЭЛИЙН ТӨСӨВ" from being detected as new section
        if 'НИЙТ' in col1_upper and ('ТӨСӨВ' in col1_upper or 'СУВГИЙН' in col1_upper):
            try:
                total_idx = column_map['total']
                total_val = row.iloc[total_idx] if total_idx < len(row) else None
                if pd.notna(total_val):
                    total_num = float(str(total_val).replace(',', '').replace(' ', ''))
                    
                    # Global totals
                    if col1_upper == 'НИЙТ ТӨСӨВ' or col1_upper == 'НИЙТ ТӨСӨВ ':
                        totals["total_budget"] = total_num
                    elif 'НИЙТ БОДИТ ТӨСӨВ' in col1_upper:
                        totals["actual_budget"] = total_num
                    elif 'ДОТООД СУРТАЛЧИЛГААНЫ СУВГИЙН ЗАРДАЛ' in col1_upper:
                        totals["internal_expense"] = total_num
                    elif 'УРТ ХУГАЦААНЫ ГЭРЭЭТ' in col1_upper and 'ЗАРДАЛ' in col1_upper:
                        totals["long_term_expense"] = total_num
                    elif current_section:
                        # Section total - save and close section
                        current_section['total'] = total_num
                        
                        if current_subsection and (current_subsection['rows'] or current_subsection['total'] > 0):
                            current_section['subsections'].append(current_subsection)
                            current_subsection = None
                        
                        if current_section['rows'] or current_section['subsections']:
                            sections.append(current_section)
                        current_section = None
            except:
                pass
            continue
        
        # Detect main section headers (but skip if contains "ЗАРДАЛ")
        is_main_section = False
        if 'ЗАРДАЛ' not in col1_upper:  # ЗАРДАЛ мөрүүд нь section header биш
            for section_name in MAIN_SECTIONS:
                if section_name in col1_upper or col1_upper == section_name:
                    # Save previous section
                    if current_section:
                        if current_subsection and (current_subsection['rows'] or current_subsection['total'] > 0):
                            current_section['subsections'].append(current_subsection)
                        if current_section['rows'] or current_section['subsections']:
                            sections.append(current_section)
                    
                    current_section = {
                        'name': col1,
                        'rows': [],
                        'total': 0,
                        'subsections': []
                    }
                    current_subsection = None
                    # Reset forward fill values for new section
                    last_owner = ""
                    last_date = ""
                    is_main_section = True
                    break
        
        if is_main_section:
            continue
        
        # Detect subsection headers
        if col1_upper in SUBSECTIONS and current_section:
            if current_subsection and (current_subsection['rows'] or current_subsection['total'] > 0):
                current_section['subsections'].append(current_subsection)
            current_subsection = {
                'name': col1,
                'rows': [],
                'total': 0
            }
            # Reset forward fill values for new subsection
            last_owner = ""
            last_date = ""
            continue
        
        # Skip ТӨРӨЛ, ХИЙГДЭХ АЖИЛ header rows
        if col1_upper in ['ТӨРӨЛ', 'ХИЙГДЭХ АЖИЛ']:
            continue
        
        # Detect subsection totals (ТӨСӨВ, ЦАЦАЛТ ТӨСӨВ, БЭЛТГЭЛ ТӨСӨВ, etc.)
        if 'ТӨСӨВ' in col1_upper and 'НИЙТ' not in col1_upper:
            try:
                total_idx = column_map['total']
                total_val = row.iloc[total_idx] if total_idx < len(row) else None
                if pd.notna(total_val) and current_subsection:
                    current_subsection['total'] = float(str(total_val).replace(',', '').replace(' ', ''))
            except:
                pass
            continue
        
        # Data rows (start with number)
        if col1.isdigit() and current_section:
            # Get values using column map
            def get_val(idx):
                if idx is None or idx >= len(row_vals):
                    return ""
                return row_vals[idx]
            
            # Get total from correct column
            try:
                total_idx = column_map['total']
                budget_val = float(str(row.iloc[total_idx]).replace(',', '')) if total_idx < len(row) and pd.notna(row.iloc[total_idx]) else 0
            except:
                budget_val = 0
            
            # Get note from correct column
            note_idx = column_map['note']
            note_val = get_val(note_idx) if note_idx else ""
            
            # Get owner with forward fill (for merged cells)
            owner_val = get_val(column_map['owner'])
            if owner_val:
                last_owner = owner_val
            else:
                owner_val = last_owner
            
            # Get date with forward fill (for merged cells)
            date_val = ""
            if column_map['date']:
                date_val = get_val(column_map['date'])
                if date_val:
                    last_date = date_val
                else:
                    date_val = last_date
            
            row_data = {
                'no': col1,
                'type': get_val(column_map['no'] + 1),  # Type is usually next to №
                'task': get_val(column_map['no'] + 2),  # Task is next
                'vendor': get_val(column_map['no'] + 3),  # Vendor/Influencer
                'date': date_val,
                'owner': owner_val,
                'freq': get_val(column_map['freq']),
                'unit_price': get_val(column_map['unit_price']),
                'total': budget_val,
                'note': note_val
            }
            
            if current_subsection:
                current_subsection['rows'].append(row_data)
            else:
                current_section['rows'].append(row_data)
    
    # Save last section
    if current_subsection and current_section:
        current_section['subsections'].append(current_subsection)
    if current_section and (current_section['rows'] or current_section['subsections']):
        sections.append(current_section)
    
    return sections, totals


def get_template_sheet(excel_path: str):
    """Read TEMPLATE sheet from Excel file."""
    try:
        xl = pd.ExcelFile(excel_path)
        
        # Find TEMPLATE sheet
        target_sheet = None
        for sn in xl.sheet_names:
            if sn.upper() == 'TEMPLATE':
                target_sheet = sn
                break
        
        if target_sheet is None:
            for sn in xl.sheet_names:
                if 'гүйцэтгэл' in sn.lower():
                    target_sheet = sn
                    break
        
        if target_sheet is None:
            return None, "TEMPLATE sheet олдсонгүй"
        
        df = pd.read_excel(xl, sheet_name=target_sheet, header=None)
        return df, target_sheet
    except Exception as e:
        return None, str(e)


def display_budget_header(header: dict):
    """Display budget header information."""
    # Row 1: Approver info
    col1, col2, col3 = st.columns([1, 2, 2])
    with col1:
        st.markdown(f"**{header['approver']}**")
    with col2:
        st.markdown(f"**{header['approver_name']}**")
    with col3:
        st.markdown(f"*{header['approver_title']}*")
    
    st.divider()
    
    # Campaign name
    st.markdown(f"### 📋 {header['campaign_name']}")
    
    # Info grid
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown("**📅 Огноо**")
        st.write(header['date'])
    with col2:
        st.markdown("**🏷️ Төсвийн код**")
        st.write(header['budget_code'])
    with col3:
        st.markdown("**📌 Маркетингийн код**")
        st.write(header['marketing_code'])
    with col4:
        st.markdown("**🏢 Компани**")
        st.write(header['company'])
    
    # Goals and targets
    if header['goal']:
        st.markdown(f"**🎯 Зорилго:** {header['goal']}")
    if header['target_audience']:
        st.markdown(f"**👥 Зорилтот хэрэглэгчид:** {header['target_audience']}")
    if header['main_message']:
        st.markdown(f"**💬 Голлох мессеж:** {header['main_message']}")
    if header['target']:
        st.markdown(f"**🎯 Таргет:** {header['target']}")
    if header['period']:
        st.markdown(f"**📆 Хугацаа:** {header['period']}")
    
    # Distribution
    col1, col2, col3 = st.columns(3)
    with col1:
        if header['content_production']:
            st.metric("Content Production", header['content_production'])
    with col2:
        if header['channel_distribution']:
            st.metric("Channel Distribution", header['channel_distribution'])
    with col3:
        if header['event_level']:
            st.write(f"**Түвшин:** {header['event_level']}")
        if header['scope']:
            st.write(f"**Хамрах хүрээ:** {header['scope']}")


def display_sections(sections: list, totals: dict, show_details: bool = True):
    """Display parsed sections with summary and details."""
    if not sections:
        st.warning("⚠️ Сувгийн мэдээлэл олдсонгүй")
        return 0
    
    st.markdown("### 📊 МАРКЕТИНГИЙН ТӨСӨВ")
    
    # Summary cards
    num_sections = len(sections)
    cols = st.columns(min(num_sections, 5))
    
    for i, section in enumerate(sections):
        col_idx = i % 5
        with cols[col_idx]:
            # Short name for display
            short_name = section['name']
            short_name = short_name.replace('СУРТАЛЧИЛГААНЫ СУВАГ', '').replace('СУВАГ', '').strip()
            if not short_name:
                short_name = section['name'][:12]
            st.metric(short_name, f"₮{section['total']:,.0f}")
    
    # Global totals
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("💰 НИЙТ ТӨСӨВ", f"₮{totals['total_budget']:,.0f}")
    with col2:
        st.metric("🏠 Дотоод зардал", f"₮{totals['internal_expense']:,.0f}")
    with col3:
        # Төсвийн зөрүү = Нийт төсөв - Нийт бодит төсөв
        budget_diff = totals['total_budget'] - totals['actual_budget']
        diff_color = "normal" if budget_diff >= 0 else "inverse"
        st.metric("📊 Төсвийн зөрүү", f"₮{budget_diff:,.0f}", delta=f"₮{budget_diff:,.0f}", delta_color=diff_color)
    with col4:
        st.metric("✅ НИЙТ БОДИТ ТӨСӨВ", f"₮{totals['actual_budget']:,.0f}")
    
    if show_details and sections:
        st.divider()
        st.markdown("### 📑 Сувгийн дэлгэрэнгүй")
        
        # Create tabs for each section
        tab_names = []
        for s in sections:
            name = s['name'].replace('СУРТАЛЧИЛГААНЫ СУВАГ', '').replace('СУВАГ', '').strip()
            if not name:
                name = s['name'][:12]
            tab_names.append(name)
        
        tabs = st.tabs(tab_names)
        
        for idx, tab in enumerate(tabs):
            section = sections[idx]
            with tab:
                st.markdown(f"#### {section['name']}")
                
                if section['total'] > 0:
                    st.info(f"💰 Сувгийн нийт: ₮{section['total']:,.0f}")
                
                # Show subsections
                if section['subsections']:
                    for sub in section['subsections']:
                        st.markdown(f"**📌 {sub['name']}**")
                        
                        if sub['rows']:
                            data = []
                            for r in sub['rows']:
                                row_dict = {
                                    "№": r['no'],
                                    "Төрөл": r['type'],
                                    "Ажил": r['task'],
                                    "Гүйцэтгэгч": r['vendor'],
                                }
                                if r.get('date'):
                                    row_dict["Хугацаа"] = r['date']
                                row_dict["Хариуцагч"] = r['owner']
                                row_dict["Давтамж"] = r['freq']
                                row_dict["Нэгж үнэ"] = r['unit_price']
                                row_dict["Нийт төсөв"] = f"₮{r['total']:,.0f}" if r['total'] else ""
                                row_dict["Тайлбар"] = r['note']
                                data.append(row_dict)
                            st.dataframe(pd.DataFrame(data), width='stretch', hide_index=True)
                        
                        # Show subsection total
                        if sub['total'] > 0:
                            st.success(f"**{sub['name']} ТӨСӨВ: ₮{sub['total']:,.0f}**")
                        st.markdown("---")
                
                # Show direct rows (no subsection)
                elif section['rows']:
                    data = []
                    for r in section['rows']:
                        row_dict = {
                            "№": r['no'],
                            "Төрөл": r['type'],
                            "Ажил": r['task'],
                            "Гүйцэтгэгч": r['vendor'],
                        }
                        if r.get('date'):
                            row_dict["Хугацаа"] = r['date']
                        row_dict["Хариуцагч"] = r['owner']
                        row_dict["Давтамж"] = r['freq']
                        row_dict["Нэгж үнэ"] = r['unit_price']
                        row_dict["Нийт төсөв"] = f"₮{r['total']:,.0f}" if r['total'] else ""
                        row_dict["Тайлбар"] = r['note']
                        data.append(row_dict)
                    st.dataframe(pd.DataFrame(data), width='stretch', hide_index=True)
    
    return totals['total_budget']


def render_budget_report_tab(session):
    """
    Render the Budget Report tab with two views:
    1. Individual budget view (Төсөв тус бүрээр)
    2. Bulk view - all budgets combined (Нийт Bulk)
    """
    st.subheader("📋 Төсвийн тайлан")
    
    # Load all files from database
    statement = (
        select(BudgetFile, User)
        .outerjoin(User, BudgetFile.uploader_id == User.id)
        .order_by(BudgetFile.uploaded_at.desc())
    )
    results = session.exec(statement).all()
    
    if not results:
        st.warning("📭 Одоогоор системд төсөв байхгүй байна.")
        st.page_link("pages/2_📤_Upload.py", label="📤 Төсөв оруулах", icon="📤")
        return
    
    files = [r[0] for r in results]
    
    # =========================================================================
    # Two Main Views: Individual vs Bulk
    # =========================================================================
    view_tab1, view_tab2 = st.tabs(["📄 Төсөв тус бүрээр", "📊 Нийт (Bulk)"])
    
    # =========================================================================
    # TAB 1: Individual Budget View
    # =========================================================================
    with view_tab1:
        file_options = {f"{f.campaign_name or f.budget_code or f.filename}": f.id for f in files}
        selected_option = st.selectbox(
            "📁 Төсөв сонгох:",
            options=list(file_options.keys()),
            key="individual_file_select"
        )
        
        if not selected_option:
            return
        
        selected_file_id = file_options[selected_option]
        selected_file = next((f for f in files if f.id == selected_file_id), None)
        
        if not selected_file:
            return
        
        # Read and display Excel
        excel_path = get_excel_file_path(selected_file.id)
        
        if not excel_path or not os.path.exists(excel_path):
            st.warning("⚠️ Excel файл олдсонгүй")
            return
        
        df, sheet_name = get_template_sheet(excel_path)
        
        if df is None:
            st.error(f"❌ {sheet_name}")
            return
        
        st.success(f"📄 Sheet: **{sheet_name}**")
        
        # Parse header and sections
        header = parse_budget_header(df)
        sections, totals = parse_excel_sections(df)
        
        # Display header
        display_budget_header(header)
        
        st.divider()
        
        # Display sections
        display_sections(sections, totals, show_details=True)
        
        # Download button
        st.divider()
        with open(excel_path, 'rb') as f:
            st.download_button(
                label="📥 Excel татах",
                data=f.read(),
                file_name=selected_file.filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="individual_download"
            )
    
    # =========================================================================
    # TAB 2: Bulk View - All Budgets Combined
    # =========================================================================
    with view_tab2:
        st.markdown("### 📊 Бүх төсвүүдийн нэгдсэн харагдац")
        
        # Collect all data from files
        all_files_data = []
        all_channel_totals = {}
        all_channel_details = {}
        grand_totals = {"total_budget": 0, "actual_budget": 0}
        
        for f in files:
            excel_path = get_excel_file_path(f.id)
            if not excel_path or not os.path.exists(excel_path):
                continue
            
            df, sheet_name = get_template_sheet(excel_path)
            if df is None:
                continue
            
            header = parse_budget_header(df)
            sections, totals = parse_excel_sections(df)
            
            all_files_data.append({
                'file': f,
                'header': header,
                'sections': sections,
                'totals': totals
            })
            
            grand_totals["total_budget"] += totals["total_budget"]
            grand_totals["actual_budget"] += totals["actual_budget"]
            
            # Aggregate by channel
            for section in sections:
                ch_name = section['name'].replace('СУРТАЛЧИЛГААНЫ СУВАГ', '').replace('СУВАГ', '').strip()
                if not ch_name:
                    ch_name = section['name'][:20]
                
                if ch_name not in all_channel_totals:
                    all_channel_totals[ch_name] = 0
                    all_channel_details[ch_name] = []
                
                all_channel_totals[ch_name] += section['total']
                all_channel_details[ch_name].append({
                    'file_name': header['campaign_name'] or f.budget_code or f.filename,
                    'section': section
                })
        
        # Summary metrics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("📁 Нийт төсвийн тоо", len(all_files_data))
        with col2:
            st.metric("💰 Нийт төсөв", f"₮{grand_totals['total_budget']:,.0f}")
        with col3:
            st.metric("✅ Нийт бодит", f"₮{grand_totals['actual_budget']:,.0f}")
        
        st.divider()
        
        # All files summary table
        st.markdown("### 📋 Бүх төсвүүдийн жагсаалт")
        
        summary_data = []
        for fd in all_files_data:
            header = fd['header']
            totals = fd['totals']
            f = fd['file']
            
            summary_data.append({
                "Кампанит ажил": header['campaign_name'] or "-",
                "Төсвийн код": header['budget_code'] or f.budget_code or "-",
                "Компани": header['company'] or "-",
                "Огноо": header['date'] or "-",
                "Хугацаа": header['period'] or "-",
                "Нийт төсөв": f"₮{totals['total_budget']:,.0f}",
                "Бодит төсөв": f"₮{totals['actual_budget']:,.0f}"
            })
        
        if summary_data:
            st.dataframe(pd.DataFrame(summary_data), width='stretch', hide_index=True)
        
        st.divider()
        
        # Channel aggregation
        st.markdown("### 📑 Сувгаар нэгтгэсэн төсөв")
        
        if all_channel_totals:
            channels_list = sorted(all_channel_totals.items(), key=lambda x: x[1], reverse=True)
            num_channels = len(channels_list)
            cols = st.columns(min(num_channels, 5))
            
            for i, (ch_name, ch_total) in enumerate(channels_list):
                col_idx = i % 5
                with cols[col_idx]:
                    num_files = len(all_channel_details[ch_name])
                    st.metric(ch_name, f"₮{ch_total:,.0f}", f"{num_files} файл")
            
            st.markdown("---")
            st.metric("🏆 НИЙТ СУВГИЙН ТӨСӨВ", f"₮{sum(all_channel_totals.values()):,.0f}")
            
            st.divider()
            
            # Detailed tabs by channel
            st.markdown("### 📊 Сувгийн дэлгэрэнгүй")
            
            channel_tabs = st.tabs([ch[0] for ch in channels_list])
            
            for idx, tab in enumerate(channel_tabs):
                ch_name, ch_total = channels_list[idx]
                with tab:
                    st.markdown(f"#### {ch_name}")
                    st.info(f"💰 Нийт: ₮{ch_total:,.0f} ({len(all_channel_details[ch_name])} файлаас)")
                    
                    for file_data in all_channel_details[ch_name]:
                        section = file_data['section']
                        
                        with st.expander(f"📁 {file_data['file_name']} - ₮{section['total']:,.0f}", expanded=False):
                            if section['subsections']:
                                for sub in section['subsections']:
                                    st.markdown(f"**📌 {sub['name']}** (₮{sub['total']:,.0f})")
                                    
                                    if sub['rows']:
                                        data = []
                                        for r in sub['rows']:
                                            row_dict = {
                                                "№": r['no'],
                                                "Төрөл": r['type'],
                                                "Ажил": r['task'],
                                                "Гүйцэтгэгч": r['vendor'],
                                            }
                                            if r.get('date'):
                                                row_dict["Хугацаа"] = r['date']
                                            row_dict["Нийт"] = f"₮{r['total']:,.0f}" if r['total'] else ""
                                            row_dict["Тайлбар"] = r['note']
                                            data.append(row_dict)
                                        st.dataframe(pd.DataFrame(data), width='stretch', hide_index=True)
                            
                            elif section['rows']:
                                data = []
                                for r in section['rows']:
                                    row_dict = {
                                        "№": r['no'],
                                        "Төрөл": r['type'],
                                        "Ажил": r['task'],
                                        "Гүйцэтгэгч": r['vendor'],
                                    }
                                    if r.get('date'):
                                        row_dict["Хугацаа"] = r['date']
                                    row_dict["Нийт"] = f"₮{r['total']:,.0f}" if r['total'] else ""
                                    row_dict["Тайлбар"] = r['note']
                                    data.append(row_dict)
                                st.dataframe(pd.DataFrame(data), width='stretch', hide_index=True)
        
        # Export buttons
        st.divider()
        st.markdown("### 📥 Тайлан татах")
        
        if summary_data:
            bulk_df = pd.DataFrame(summary_data)
            csv_data = bulk_df.to_csv(index=False, encoding='utf-8-sig')
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.download_button(
                    label="📥 Төсвүүдийн жагсаалт (CSV)",
                    data=csv_data,
                    file_name="budget_summary_bulk.csv",
                    mime="text/csv"
                )
            
            if all_channel_totals:
                channel_summary = [{
                    "Суваг": ch_name,
                    "Нийт төсөв": ch_total,
                    "Файлын тоо": len(all_channel_details[ch_name])
                } for ch_name, ch_total in sorted(all_channel_totals.items(), key=lambda x: x[1], reverse=True)]
                
                channel_df = pd.DataFrame(channel_summary)
                channel_csv = channel_df.to_csv(index=False, encoding='utf-8-sig')
                
                with col2:
                    st.download_button(
                        label="📥 Сувгаар нэгтгэсэн (CSV)",
                        data=channel_csv,
                        file_name="channel_summary_bulk.csv",
                        mime="text/csv"
                    )
            
            # BULK XLSX Export - All data in one Excel file with multiple sheets
            with col3:
                try:
                    xlsx_data = generate_bulk_excel_export(all_files_data, all_channel_totals, all_channel_details, grand_totals)
                    if xlsx_data:
                        date_str = datetime.now().strftime("%Y%m%d_%H%M")
                        st.download_button(
                            label="📥 Нэгтгэсэн XLSX татах",
                            data=xlsx_data,
                            file_name=f"Budget_Bulk_Export_{date_str}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                except Exception as e:
                    st.error(f"XLSX үүсгэхэд алдаа: {e}")


# =============================================================================
# EXPORT TAB
# =============================================================================

def render_export_tab(session):
    """Render the PDF Export tab."""
    
    # Load files for export - EXCLUDE REJECTED files
    statement = (
        select(BudgetFile)
        .where(BudgetFile.status != FileStatus.REJECTED)
        .order_by(BudgetFile.uploaded_at.desc())
    )
    files = list(session.exec(statement).all())
    
    if not files:
        st.info("Экспортлох файл байхгүй байна.")
        return
    
    st.subheader("📄 PDF Экспорт")
    
    file_data_list = []
    for file in files:
        budget_code = getattr(file, 'budget_code', None) or f"#{file.id}"
        
        # Get stored file path
        excel_path = None
        if file.pdf_file_path:
            # Check for original Excel in uploaded_files
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            possible_excel = os.path.join(base_dir, 'assets', 'uploaded_files', f"budget_{file.id}_*.xlsx")
            import glob
            matches = glob.glob(possible_excel.replace("*", "*"))
            if matches:
                excel_path = matches[0]
        
        file_data_list.append({
            "id": file.id,
            "budget_code": budget_code,
            "filename": file.filename,
            "total_amount": f"₮{float(file.total_amount):,.0f}" if file.total_amount else "N/A",
            "excel_path": excel_path
        })
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        selected_codes = [f"{fd['budget_code']} - {fd['filename']}" for fd in file_data_list]
        selected = st.selectbox("Файл сонгох", selected_codes)
    
    if selected:
        selected_idx = selected_codes.index(selected)
        selected_file = file_data_list[selected_idx]
        
        with col2:
            st.write("")
            excel_path = selected_file.get('excel_path')
            if excel_path and os.path.exists(excel_path):
                # Excel download button
                with open(excel_path, 'rb') as f:
                    excel_bytes = f.read()
                st.download_button(
                    label="📥 Excel татах",
                    data=excel_bytes,
                    file_name=selected_file['filename'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                
                # PDF download button
                pdf_bytes = generate_excel_pdf(selected_file, excel_path)
                if pdf_bytes:
                    st.download_button(
                        label="📥 PDF татах",
                        data=pdf_bytes,
                        file_name=f"budget_{selected_file['budget_code'].replace('#', '').replace('/', '_')}.pdf",
                        mime="application/pdf"
                    )
            else:
                st.warning("Excel файл олдсонгүй")
        
        # Show file details
        st.divider()
        st.subheader("📋 Сонгосон файлын дэлгэрэнгүй")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Төсвийн код", selected_file['budget_code'])
        with col2:
            st.metric("Нийт дүн", selected_file['total_amount'])
        with col3:
            st.metric("Файлын нэр", selected_file['filename'][:30] + "..." if len(selected_file['filename']) > 30 else selected_file['filename'])


# =============================================================================
# CPP REPORT TAB - EDITABLE GRID WITH DATABASE SAVE
# =============================================================================

def get_cpp_items_for_sheet(sheet_name: str):
    """Get CPP items from database for a specific sheet."""
    from database import get_session, CppBudgetItem
    from sqlmodel import select
    
    with get_session() as session:
        items = session.exec(
            select(CppBudgetItem)
            .where(CppBudgetItem.category_name == sheet_name)
            .order_by(CppBudgetItem.owner_username, CppBudgetItem.row_number)
        ).all()
        
        # Convert to list of dicts for DataFrame
        data = []
        for item in items:
            custom = item.get_custom_fields()
            row = {
                '_id': item.id,
                '_owner_id': item.owner_id,
                '_owner': item.owner_username,
                **custom
            }
            data.append(row)
        
        return data


def save_cpp_items_for_sheet(sheet_name: str, owner_id: int, owner_username: str, rows: list):
    """Save CPP items to database for a specific sheet."""
    from database import get_session, CppBudgetItem
    from sqlmodel import select
    import json
    from datetime import datetime
    
    with get_session() as session:
        # Delete existing rows for this user & sheet
        existing = session.exec(
            select(CppBudgetItem)
            .where(CppBudgetItem.category_name == sheet_name)
            .where(CppBudgetItem.owner_id == owner_id)
        ).all()
        
        for item in existing:
            session.delete(item)
        
        # Add new rows
        for idx, row_data in enumerate(rows, 1):
            # Filter out internal columns
            clean_data = {k: v for k, v in row_data.items() if not k.startswith('_') and pd.notna(v)}
            
            new_item = CppBudgetItem(
                owner_id=owner_id,
                owner_username=owner_username,
                category_name=sheet_name,
                row_number=idx,
                custom_fields=json.dumps(clean_data, ensure_ascii=False, default=str),
                status="draft",
                updated_at=datetime.utcnow()
            )
            session.add(new_item)
        
        session.commit()


def get_tv_data_from_budget_files(user_id: int = None):
    """
    Extract TV section data from all budget files for CPP report.
    Returns list of rows in CPP format.
    """
    from database import get_session, BudgetFile
    from sqlmodel import select
    import os
    from config import UPLOAD_FOLDER
    
    # Map budget code first letter to company name
    company_map = {
        'A': 'Юнител',
        'B': 'Юнивишн',
        'G': 'Green Future',
        'J': 'IVLBS',
        'T': 'MPSC'
    }
    
    tv_rows = []
    row_num = 1
    
    with get_session() as session:
        # Get all budget files (or filter by user if needed)
        query = select(BudgetFile)
        if user_id:
            query = query.where(BudgetFile.uploader_id == user_id)
        
        files = session.exec(query).all()
        
        for bf in files:
            # Get file path
            file_path = os.path.join(UPLOAD_FOLDER, bf.filename) if bf.filename else None
            if not file_path or not os.path.exists(file_path):
                continue
            
            try:
                # Read and parse the Excel file
                df, sheet_name = get_template_sheet(file_path)
                if df is None:
                    continue
                
                sections, _ = parse_excel_sections(df)
                
                # Debug: Check sections structure
                if not isinstance(sections, list):
                    continue
                
                # Find TV sections
                tv_sections = []
                for s in sections:
                    if not isinstance(s, dict):
                        continue
                    section_name = s.get('name', '')
                    if not isinstance(section_name, str):
                        continue
                    if any(kw in section_name.upper() for kw in ['ТВ СУВАГ', 'ТВ СУРТАЛЧИЛГАА', 'TV']):
                        tv_sections.append(s)
                
                # Get company from budget_code
                company = ''
                if bf.budget_code and len(bf.budget_code) > 0:
                    company_code = bf.budget_code[0].upper()
                    company = company_map.get(company_code, '')
                
                for section in tv_sections:
                    # Process section rows
                    for row in section.get('rows', []):
                        if not isinstance(row, dict):
                            continue
                        tv_row = {
                            '№': row_num,
                            'Кампанит ажлын нэр': bf.campaign_name or '',
                            'Брэнд': bf.brand or '',
                            'Компани': company,
                            'Мэргэжилтэн': bf.specialist_name or '',
                            'Төсвийн код': bf.budget_code or '',
                            'Үнийн дүн (₮)': row.get('total', ''),
                            'Төсвийн эх үүсвэр': 'Кампанит ажил',
                            'Сурталчилгааны төрөл': row.get('task', ''),
                            'Сувгийн нэр': row.get('type', ''),
                            'Эхлэх өдөр': None,
                            'Дуусах өдөр': None,
                            'Нийт өдөр': None,
                            'Сурталчилгааны цаг': '',
                            'Сурталчилгааны урт': row.get('unit_price', ''),  # Metric
                            'Нийт давтамж': row.get('freq', ''),
                            'Нийт сурталчилгааны урт': None,
                            '1 секундын үнэлгээ (₮)': '',
                            'Тайлбар': row.get('note', ''),
                            '_source_file': bf.filename,
                            '_source_campaign': bf.campaign_name,
                        }
                        tv_rows.append(tv_row)
                        row_num += 1
                    
                    # Process subsection rows
                    for sub in section.get('subsections', []):
                        if not isinstance(sub, dict):
                            continue
                        for row in sub.get('rows', []):
                            if not isinstance(row, dict):
                                continue
                            tv_row = {
                                '№': row_num,
                                'Кампанит ажлын нэр': bf.campaign_name or '',
                                'Брэнд': bf.brand or '',
                                'Компани': company,
                                'Мэргэжилтэн': bf.specialist_name or '',
                                'Төсвийн код': bf.budget_code or '',
                                'Үнийн дүн (₮)': row.get('total', ''),
                                'Төсвийн эх үүсвэр': 'Кампанит ажил',
                                'Сурталчилгааны төрөл': sub.get('name', '') or row.get('task', ''),
                                'Сувгийн нэр': row.get('type', ''),
                                'Эхлэх өдөр': None,
                                'Дуусах өдөр': None,
                                'Нийт өдөр': None,
                                'Сурталчилгааны цаг': '',
                                'Сурталчилгааны урт': row.get('unit_price', ''),
                                'Нийт давтамж': row.get('freq', ''),
                                'Нийт сурталчилгааны урт': None,
                                '1 секундын үнэлгээ (₮)': '',
                                'Тайлбар': row.get('note', ''),
                                '_source_file': bf.filename,
                                '_source_campaign': bf.campaign_name,
                            }
                            tv_rows.append(tv_row)
                            row_num += 1
            except Exception as e:
                import traceback
                st.warning(f"Excel татах алдаа ({bf.filename}): {str(e)}")
                continue
    
    return tv_rows


def get_campaign_options():
    """Get list of campaign names from budget files for dropdown."""
    from database import get_session, BudgetFile, BudgetItem
    from sqlmodel import select
    
    # Map budget code first letter to company name
    company_map = {
        'A': 'Юнител',
        'B': 'Юнивишн',
        'G': 'Green Future',
        'J': 'IVLBS',
        'T': 'MPSC'
    }
    
    with get_session() as session:
        # Get all budget files with campaign names
        files = session.exec(select(BudgetFile)).all()
        
        campaigns = []
        seen_names = set()
        for f in files:
            if f.campaign_name and f.campaign_name not in seen_names:
                seen_names.add(f.campaign_name)
                
                # Get company from budget_code first letter
                company = ''
                if f.budget_code and len(f.budget_code) > 0:
                    company_code = f.budget_code[0].upper()
                    company = company_map.get(company_code, '')
                
                # Fallback: try to get from BudgetItem vendor field
                if not company:
                    first_item = session.exec(
                        select(BudgetItem).where(BudgetItem.file_id == f.id)
                    ).first()
                    if first_item and first_item.vendor:
                        company = first_item.vendor
                
                campaigns.append({
                    'name': f.campaign_name,
                    'budget_code': f.budget_code or '',
                    'company': company,
                    'brand': f.brand or '',
                    'specialist': f.specialist_name or '',
                })
        return campaigns


def get_budget_codes_for_campaign(campaign_name: str):
    """Get budget codes for a specific campaign."""
    from database import get_session, BudgetFile
    from sqlmodel import select
    
    with get_session() as session:
        files = session.exec(
            select(BudgetFile).where(BudgetFile.campaign_name == campaign_name)
        ).all()
        return [f.budget_code for f in files if f.budget_code]


def render_tv_sheet(current_user):
    """
    Render TV ads sheet with specialized dropdowns, dates, and formulas.
    Auto-imports TV data from budget files.
    """
    from modules.report_generator import (
        TV_ADS_COLUMNS, TV_BUDGET_SOURCE_OPTIONS, 
        TV_AD_TYPE_OPTIONS, TV_CHANNEL_OPTIONS
    )
    
    current_user_id = current_user.id if current_user else None
    current_username = current_user.username if current_user else None
    current_fullname = current_user.full_name if current_user else current_username
    
    # Get campaign options for dropdown
    campaigns = get_campaign_options()
    campaign_names = [''] + [c['name'] for c in campaigns]
    campaign_dict = {c['name']: c for c in campaigns}
    
    # === AUTO-IMPORT TV DATA FROM BUDGET FILES ===
    st.markdown("### 📺 TV сурталчилгааны мэдээлэл")
    
    # Button to refresh/import data
    col_refresh, col_info = st.columns([1, 3])
    with col_refresh:
        refresh_clicked = st.button("🔄 Төсвөөс татах", key="refresh_tv_data", 
                                   help="Бүх төсвийн файлуудаас TV хэсгийн мэдээллийг автоматаар татна")
    with col_info:
        st.info("💡 Таны оруулсан бүх төсвийн файлуудаас TV сувгийн мэдээлэл автоматаар гарч ирнэ")
    
    # Get TV data from budget files
    budget_tv_data = get_tv_data_from_budget_files()
    
    # Get existing CPP saved data
    db_items = get_cpp_items_for_sheet("TV ads")
    
    if not current_user:
        st.warning("⚠️ Нэвтэрч орсноор мэдээлэл засварлах боломжтой болно")
        if budget_tv_data:
            display_df = pd.DataFrame(budget_tv_data)
            # Remove internal columns for display
            display_cols = [c for c in TV_ADS_COLUMNS if c in display_df.columns]
            st.dataframe(display_df[display_cols], width='stretch', hide_index=True)
        return
    
    # Merge budget data with saved edits
    # Priority: saved data > budget data
    merged_data = []
    
    # Create lookup for saved items by source file + campaign
    saved_lookup = {}
    for item in db_items:
        key = (item.get('_source_file', ''), item.get('_source_campaign', ''), item.get('№', 0))
        saved_lookup[key] = item
    
    # Use budget data as base, overlay with saved edits
    row_num = 1
    for brow in budget_tv_data:
        key = (brow.get('_source_file', ''), brow.get('_source_campaign', ''), brow.get('№', 0))
        if key in saved_lookup:
            # Use saved version (user edited)
            saved_row = saved_lookup[key].copy()
            saved_row['№'] = row_num
            merged_data.append(saved_row)
        else:
            # Use budget data (auto-imported)
            brow['№'] = row_num
            merged_data.append(brow)
        row_num += 1
    
    # Add any additional saved rows not in budget (manually added)
    for item in db_items:
        if not item.get('_source_file'):  # Manually added row
            item['№'] = row_num
            merged_data.append(item)
            row_num += 1
    
    # Separate user's rows and others
    user_items = [r for r in merged_data if r.get('_owner_id') == current_user_id or not r.get('_owner_id')]
    other_items = [r for r in merged_data if r.get('_owner_id') and r.get('_owner_id') != current_user_id]
    
    # Show other users' data as read-only
    if other_items:
        with st.expander(f"🔒 Бусад хэрэглэгчдийн оруулсан ({len(other_items)} мөр)", expanded=False):
            other_df = pd.DataFrame(other_items)
            display_cols = ['_owner'] + [c for c in TV_ADS_COLUMNS if c in other_df.columns]
            if '_owner' in other_df.columns:
                other_df = other_df.rename(columns={'_owner': 'Оруулсан'})
                display_cols[0] = 'Оруулсан'
            st.dataframe(other_df[[c for c in display_cols if c in other_df.columns]], 
                        width='stretch', hide_index=True, height=200)
    
    st.divider()
    st.markdown(f"### ✏️ TV Мэдээлэл засварлах ({len(user_items)} мөр)")
    
    # === ADD NEW ROW SECTION ===
    with st.expander("➕ Гараар шинэ мөр нэмэх", expanded=False):
        add_col1, add_col2 = st.columns([3, 1])
        
        with add_col1:
            selected_campaign = st.selectbox(
                "Кампанит ажлын нэр сонгоно уу",
                options=[''] + campaign_names[1:],
                key="tv_add_campaign",
                help="Кампанит ажлын нэр сонгоход автоматаар бөглөгдөх талбарууд бөглөгдөнө"
            )
        
        with add_col2:
            st.write("")
            st.write("")
            if st.button("➕ Мөр нэмэх", key="add_tv_row", type="primary", disabled=not selected_campaign):
                if selected_campaign and selected_campaign in campaign_dict:
                    campaign_info = campaign_dict[selected_campaign]
                    new_row = {
                        '№': len(user_items) + 1,
                        'Кампанит ажлын нэр': selected_campaign,
                        'Брэнд': campaign_info.get('brand', ''),
                        'Компани': campaign_info.get('company', ''),
                        'Мэргэжилтэн': campaign_info.get('specialist', '') or current_fullname,
                        'Төсвийн код': campaign_info.get('budget_code', ''),
                        '_owner_id': current_user_id,
                        '_owner': current_username,
                    }
                    user_items.append(new_row)
                    save_cpp_items_for_sheet("TV ads", current_user_id, current_username, user_items)
                    st.success(f"✅ Шинэ мөр нэмэгдлээ: {selected_campaign}")
                    st.rerun()
        
        if selected_campaign and selected_campaign in campaign_dict:
            campaign_info = campaign_dict[selected_campaign]
            st.info(f"📋 **Автоматаар бөглөгдөх:** Брэнд: {campaign_info.get('brand', '-')} | Компани: {campaign_info.get('company', '-')} | Төсвийн код: {campaign_info.get('budget_code', '-')}")
    
    # Prepare DataFrame with proper dtypes for numeric columns
    numeric_columns = ['№', 'Үнийн дүн (₮)', 'Нийт өдөр', 'Сурталчилгааны урт', 
                       'Нийт давтамж', 'Нийт сурталчилгааны урт']
    
    if user_items:
        user_df = pd.DataFrame(user_items)
        for col in TV_ADS_COLUMNS:
            if col not in user_df.columns:
                user_df[col] = None
        edit_df = user_df[TV_ADS_COLUMNS].copy()
    else:
        edit_df = pd.DataFrame(columns=TV_ADS_COLUMNS)
        st.info("📭 Төсвийн файлд TV сувгийн мэдээлэл олдсонгүй. Төсвөө оруулсан эсэхээ шалгана уу.")
    
    # Ensure numeric columns have proper dtype
    for col in numeric_columns:
        if col in edit_df.columns:
            edit_df[col] = pd.to_numeric(edit_df[col], errors='coerce')
    
    # Convert date columns to datetime
    date_columns = ['Эхлэх өдөр', 'Дуусах өдөр']
    for col in date_columns:
        if col in edit_df.columns:
            edit_df[col] = pd.to_datetime(edit_df[col], errors='coerce')
    
    # Column configuration with dropdowns and dates
    column_config = {
        '№': st.column_config.NumberColumn('№', width=50, disabled=True),
        
        'Кампанит ажлын нэр': st.column_config.SelectboxColumn(
            'Кампанит ажлын нэр',
            options=campaign_names,
            width=200,
            help="Өмнө оруулсан кампаниудаас сонгоно"
        ),
        
        'Брэнд': st.column_config.TextColumn(
            'Брэнд',
            width=120,
            help="Кампанит ажлаас автоматаар гарна"
        ),
        
        'Компани': st.column_config.TextColumn(
            'Компани', 
            width=120,
            help="Кампанит ажлаас автоматаар гарна"
        ),
        
        'Мэргэжилтэн': st.column_config.TextColumn(
            'Мэргэжилтэн',
            width=130,
            disabled=True,
            default=current_fullname,
            help="Нэвтэрсэн хэрэглэгчээс автоматаар бөглөгдөнө"
        ),
        
        'Төсвийн код': st.column_config.TextColumn(
            'Төсвийн код',
            width=120,
            help="Кампанит ажлаас автоматаар эсвэл сонгоно"
        ),
        
        'Үнийн дүн (₮)': st.column_config.NumberColumn(
            'Үнийн дүн (₮)',
            width=140,
            disabled=True,
            format="%,.0f",
            help="= Нийт сурт. урт × 1 сек үнэлгээ (автомат)"
        ),
        
        'Төсвийн эх үүсвэр': st.column_config.SelectboxColumn(
            'Төсвийн эх үүсвэр',
            options=TV_BUDGET_SOURCE_OPTIONS,
            width=170,
            help="Кампанит ажил / Ивээн тэтгэлэг / Урт хугацааны суваг"
        ),
        
        'Сурталчилгааны төрөл': st.column_config.SelectboxColumn(
            'Сурталчилгааны төрөл',
            options=TV_AD_TYPE_OPTIONS,
            width=180,
            help="Аман сурталчилгаа / Задгай цацалт / Нэвтрүүлэг"
        ),
        
        'Сувгийн нэр': st.column_config.SelectboxColumn(
            'Сувгийн нэр',
            options=TV_CHANNEL_OPTIONS,
            width=140,
            help="TV сувгуудаас сонгоно"
        ),
        
        'Эхлэх өдөр': st.column_config.DateColumn(
            'Эхлэх өдөр',
            width=120,
            help="Сурталчилгаа эхлэх огноо"
        ),
        
        'Дуусах өдөр': st.column_config.DateColumn(
            'Дуусах өдөр',
            width=120,
            help="Сурталчилгаа дуусах огноо"
        ),
        
        'Нийт өдөр': st.column_config.NumberColumn(
            'Нийт өдөр',
            width=90,
            disabled=True,
            help="= Дуусах өдөр - Эхлэх өдөр (автомат)"
        ),
        
        'Сурталчилгааны цаг': st.column_config.TextColumn(
            'Сурталчилгааны цаг',
            width=170,
            help="Жишээ: 19:00-23:00"
        ),
        
        'Сурталчилгааны урт': st.column_config.NumberColumn(
            'Сурталчилгааны урт',
            width=160,
            help="Секундээр (тоо)"
        ),
        
        'Нийт давтамж': st.column_config.NumberColumn(
            'Нийт давтамж',
            width=120,
            help="Нийт хэдэн удаа гарах (тоо)"
        ),
        
        'Нийт сурталчилгааны урт': st.column_config.NumberColumn(
            'Нийт сурталчилгааны урт',
            width=190,
            disabled=True,
            help="= Сурт. урт × Давтамж (автомат)"
        ),
        
        '1 секундын үнэлгээ (₮)': st.column_config.TextColumn(
            '1 секундын үнэлгээ (₮)',
            width=170,
            help="1 секундын үнэ (тоо оруулна уу)"
        ),
        
        'Тайлбар': st.column_config.TextColumn(
            'Тайлбар',
            width=200,
            help="Нэмэлт тайлбар"
        ),
    }
    
    # Editable grid
    edited_df = st.data_editor(
        edit_df,
        width='stretch',
        hide_index=True,
        height=400,
        num_rows="dynamic",
        column_config=column_config,
        key=f"editor_tv_{current_user_id}"
    )
    
    # Action buttons
    col1, col2, col3, col4 = st.columns([1, 1, 1, 2])
    
    with col1:
        if st.button("💾 Хадгалах", key="save_tv", type="primary", width='stretch'):
            rows_to_save = []
            for idx, row in edited_df.iterrows():
                row_dict = row.to_dict()
                
                # Auto-fill from campaign selection
                campaign_name = row_dict.get('Кампанит ажлын нэр')
                if campaign_name and campaign_name in campaign_dict:
                    campaign_info = campaign_dict[campaign_name]
                    if not row_dict.get('Брэнд') or pd.isna(row_dict.get('Брэнд')):
                        row_dict['Брэнд'] = campaign_info.get('brand', '')
                    if not row_dict.get('Компани') or pd.isna(row_dict.get('Компани')):
                        row_dict['Компани'] = campaign_info.get('company', '')
                    if not row_dict.get('Төсвийн код') or pd.isna(row_dict.get('Төсвийн код')):
                        row_dict['Төсвийн код'] = campaign_info.get('budget_code', '')
                    # Мэргэжилтэн - кампаниас эсвэл нэвтэрсэн хэрэглэгчээс
                    if not row_dict.get('Мэргэжилтэн') or pd.isna(row_dict.get('Мэргэжилтэн')):
                        row_dict['Мэргэжилтэн'] = campaign_info.get('specialist', '') or current_fullname
                else:
                    # No campaign selected - use current user as specialist
                    if not row_dict.get('Мэргэжилтэн') or pd.isna(row_dict.get('Мэргэжилтэн')):
                        row_dict['Мэргэжилтэн'] = current_fullname
                
                # Row number
                row_dict['№'] = idx + 1
                
                # Calculate Нийт өдөр
                try:
                    start_date = pd.to_datetime(row_dict.get('Эхлэх өдөр'))
                    end_date = pd.to_datetime(row_dict.get('Дуусах өдөр'))
                    if pd.notna(start_date) and pd.notna(end_date):
                        row_dict['Нийт өдөр'] = (end_date - start_date).days
                except:
                    pass
                
                # Calculate Нийт сурталчилгааны урт
                try:
                    ad_length = float(row_dict.get('Сурталчилгааны урт') or 0)
                    frequency = float(row_dict.get('Нийт давтамж') or 0)
                    row_dict['Нийт сурталчилгааны урт'] = ad_length * frequency
                except:
                    pass
                
                # Calculate Үнийн дүн
                try:
                    total_seconds = float(row_dict.get('Нийт сурталчилгааны урт') or 0)
                    price_str = str(row_dict.get('1 секундын үнэлгээ (₮)') or '0').replace(',', '').strip()
                    price_per_sec = float(price_str) if price_str else 0
                    row_dict['Үнийн дүн (₮)'] = total_seconds * price_per_sec
                except:
                    pass
                
                # Only save non-empty rows
                if any(v for k, v in row_dict.items() if k not in ['№', '_owner', '_owner_id'] and pd.notna(v) and str(v).strip()):
                    rows_to_save.append(row_dict)
            
            save_cpp_items_for_sheet("TV ads", current_user_id, current_username, rows_to_save)
            st.success(f"✅ {len(rows_to_save)} мөр хадгалагдлаа!")
            st.rerun()
    
    with col2:
        if len(user_items) > 0:
            if st.button("🗑️ Миний мөрүүд устгах", key="clear_tv", width='stretch'):
                st.session_state["confirm_delete_tv"] = True
        else:
            st.button("🗑️ Миний мөрүүд устгах", key="clear_tv", width='stretch', disabled=True)
    
    with col3:
        if st.session_state.get("confirm_delete_tv"):
            if st.button("⚠️ Тийм, устгах", key="confirm_tv", type="secondary", width='stretch'):
                save_cpp_items_for_sheet("TV ads", current_user_id, current_username, [])
                st.session_state["confirm_delete_tv"] = False
                st.success(f"✅ TV ads дээрх таны {len(user_items)} мөр устгагдлаа!")
                st.rerun()
    
    with col4:
        total_rows = len(user_items) + len(other_items)
        st.info(f"📊 Таны мөр: {len(user_items)} | Нийт: {total_rows}")


def render_editable_sheet(sheet_name: str, current_user, columns_def: list):
    """
    Render an editable data grid for a CPP sheet.
    - Users can add, edit, delete their own rows
    - Other users' rows are read-only
    
    Args:
        sheet_name: Name of the sheet (e.g., 'TV ads')
        current_user: Current logged in user object
        columns_def: List of column names for the sheet
    """
    current_user_id = current_user.id if current_user else None
    current_username = current_user.username if current_user else None
    
    # Get data from database
    db_items = get_cpp_items_for_sheet(sheet_name)
    
    if not current_user:
        st.warning("⚠️ Нэвтэрч орсноор мэдээлэл оруулах боломжтой болно")
        # Show all data as read-only if any
        if db_items:
            all_df = pd.DataFrame(db_items)
            display_cols = [c for c in columns_def if c in all_df.columns]
            if display_cols:
                st.dataframe(all_df[display_cols], width='stretch', hide_index=True)
        return
    
    # Separate user's rows and others
    user_items = [r for r in db_items if r.get('_owner_id') == current_user_id]
    other_items = [r for r in db_items if r.get('_owner_id') != current_user_id]
    
    # Show other users' data as read-only
    if other_items:
        with st.expander(f"🔒 Бусад хэрэглэгчдийн оруулсан ({len(other_items)} мөр)", expanded=False):
            other_df = pd.DataFrame(other_items)
            # Add owner column for display
            display_cols = ['_owner'] + [c for c in columns_def if c in other_df.columns]
            other_df = other_df.rename(columns={'_owner': 'Оруулсан'})
            display_cols[0] = 'Оруулсан'
            st.dataframe(other_df[display_cols], width='stretch', hide_index=True, height=200)
    
    st.divider()
    
    # User's editable section
    st.markdown(f"### ✏️ Таны мөрүүд ({current_username})")
    
    # Prepare DataFrame for editing
    if user_items:
        user_df = pd.DataFrame(user_items)
        # Ensure all columns exist
        for col in columns_def:
            if col not in user_df.columns:
                user_df[col] = ""
        edit_df = user_df[columns_def].copy()
    else:
        # Empty DataFrame with proper columns - start with one empty row
        edit_df = pd.DataFrame(columns=columns_def)
    
    # Column configuration for better UX
    column_config = {}
    for col in columns_def:
        if col == '№':
            column_config[col] = st.column_config.NumberColumn(col, width="small")
        elif 'дүн' in col.lower() or 'зардал' in col.lower() or 'үнэ' in col.lower():
            column_config[col] = st.column_config.NumberColumn(col, format="₮%d")
        elif col.lower() in ['start', 'end', 'start date', 'end date']:
            column_config[col] = st.column_config.DateColumn(col)
        elif 'days' in col.lower() or 'давтамж' in col.lower() or 'тоо' in col.lower() or 'сек' in col.lower():
            column_config[col] = st.column_config.NumberColumn(col)
    
    # Editable grid with better styling
    edited_df = st.data_editor(
        edit_df,
        width='stretch',
        hide_index=True,
        height=400,
        num_rows="dynamic",  # Allow adding/removing rows
        column_config=column_config,
        key=f"editor_{sheet_name}_{current_user_id}"
    )
    
    # Action buttons
    col1, col2, col3, col4 = st.columns([1, 1, 1, 2])
    
    with col1:
        if st.button("💾 Хадгалах", key=f"save_{sheet_name}", type="primary", width='stretch'):
            # Convert to list of dicts
            rows_to_save = edited_df.to_dict('records')
            # Filter out completely empty rows
            rows_to_save = [r for r in rows_to_save if any(v for k, v in r.items() if pd.notna(v) and str(v).strip())]
            
            save_cpp_items_for_sheet(sheet_name, current_user_id, current_username, rows_to_save)
            st.success(f"✅ {len(rows_to_save)} мөр хадгалагдлаа!")
            st.rerun()
    
    with col2:
        # Only delete current user's rows in this sheet
        if len(user_items) > 0:
            if st.button("🗑️ Миний мөрүүд устгах", key=f"clear_{sheet_name}", width='stretch'):
                st.session_state[f"confirm_delete_{sheet_name}"] = True
        else:
            st.button("🗑️ Миний мөрүүд устгах", key=f"clear_{sheet_name}", width='stretch', disabled=True)
    
    with col3:
        if st.session_state.get(f"confirm_delete_{sheet_name}"):
            if st.button("⚠️ Тийм, устгах", key=f"confirm_{sheet_name}", type="secondary", width='stretch'):
                # Delete only current user's rows in this sheet
                save_cpp_items_for_sheet(sheet_name, current_user_id, current_username, [])
                st.session_state[f"confirm_delete_{sheet_name}"] = False
                st.success(f"✅ '{sheet_name}' дээрх таны {len(user_items)} мөр устгагдлаа!")
                st.rerun()
    
    with col4:
        total_rows = len(user_items) + len(other_items)
        st.info(f"📊 Таны мөр: {len(user_items)} | Нийт: {total_rows}")


def render_cpp_report_tab(session):
    """Render the CPP Report tab with editable grids."""
    
    st.subheader("📊 CPP Report - Гараар бөглөх")
    
    # Get current user for row-level security
    jwt_user = get_current_user_from_token()
    
    if not jwt_user:
        st.warning("⚠️ Нэвтрэх шаардлагатай")
        return
    
    # Get user from database for full object
    with get_session() as session:
        current_user = session.get(User, int(jwt_user['id']))
    
    if not current_user:
        st.warning("⚠️ Нэвтрэх шаардлагатай")
        return
    
    # Get CPP items count for summary
    from database import get_session as db_session, CppBudgetItem
    from sqlmodel import select, func
    
    with db_session() as sess:
        total_items = len(sess.exec(select(CppBudgetItem)).all())
        user_items = len(sess.exec(
            select(CppBudgetItem).where(CppBudgetItem.owner_id == current_user.id)
        ).all())
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📝 Таны мөр", f"{user_items} ш")
    with col2:
        st.metric("📊 Нийт мөр", f"{total_items} ш")
    with col3:
        st.write("")  # placeholder
    with col4:
        # Download button at top (exports CPP items)
        if total_items > 0:
            try:
                from modules.report_generator import export_cpp_items_to_excel
                excel_data = export_cpp_items_to_excel(session)
                date_str = datetime.now().strftime("%Y%m%d")
                
                st.download_button(
                    label="📥 Excel татах",
                    data=excel_data,
                    file_name=f"CPP_Report_{date_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    width='stretch'
                )
            except Exception as e:
                st.error(f"Excel татах алдаа: {e}")
    
    st.divider()
    
    # Sheet tabs for editing
    sheet_tabs = st.tabs([
        "📊 General",
        "📺 TV ads",
        "🏙️ OOH & DOOH ads",
        "🏢 Indoor ads",
        "📻 FM ads",
        "🎬 Cinema ads",
        "🛒 Shopping mall",
        "📱 Digital & Social"
    ])
    
    # General Tab - Summary of all data
    with sheet_tabs[0]:
        st.markdown("### 📊 Ерөнхий тоймлол")
        # Calculate summary from CPP items
        from database import get_session as db_session, CppBudgetItem
        from sqlmodel import select
        
        with db_session() as sess:
            all_items = sess.exec(select(CppBudgetItem)).all()
        
        if all_items:
            # Group by sheet/category
            summary_data = {}
            for item in all_items:
                cat = item.category_name or "Бусад"
                if cat not in summary_data:
                    summary_data[cat] = {"count": 0, "total": 0}
                summary_data[cat]["count"] += 1
                # Try to get total from custom fields
                custom = item.get_custom_fields()
                for key in ['Нийт дүн', 'Нийт төсөв', 'TV ads ACTUAL']:
                    if key in custom:
                        try:
                            val = float(str(custom[key]).replace(',', '').replace('₮', ''))
                            summary_data[cat]["total"] += val
                        except:
                            pass
            
            summary_df = pd.DataFrame([
                {"Категори": k, "Мөрийн тоо": v["count"], "Нийт дүн": f"₮{v['total']:,.0f}"}
                for k, v in summary_data.items()
            ])
            st.dataframe(summary_df, width='stretch', hide_index=True)
        else:
            st.info("📭 Өгөгдөл байхгүй байна. Tab бүрт мөр нэмж эхлээрэй.")
    
    # Import column definitions
    from modules.report_generator import (
        TV_ADS_COLUMNS, OOH_ADS_COLUMNS, INDOOR_ADS_COLUMNS,
        FM_ADS_COLUMNS, CINEMA_ADS_COLUMNS, SHOPPING_MALL_COLUMNS, DIGITAL_COLUMNS
    )
    
    # TV ads Tab - Uses specialized render function with dropdowns
    with sheet_tabs[1]:
        st.markdown("### 📺 ТВ Сурталчилгааны Суваг")
        st.caption("Dropdown, огноо, автомат тооцоолол бүхий ТВ сурталчилгааны өгөгдөл")
        render_tv_sheet(current_user)
    
    # OOH & DOOH ads Tab
    with sheet_tabs[2]:
        st.markdown("### 🏙️ OOH & DOOH ads")
        st.caption("Гадна самбарын сурталчилгаа")
        render_editable_sheet("OOH & DOOH ads", current_user, OOH_ADS_COLUMNS)
    
    # Indoor ads Tab
    with sheet_tabs[3]:
        st.markdown("### 🏢 Indoor ads")
        st.caption("Дотоод сурталчилгаа")
        render_editable_sheet("Indoor ads", current_user, INDOOR_ADS_COLUMNS)
    
    # FM ads Tab
    with sheet_tabs[4]:
        st.markdown("### 📻 FM ads")
        st.caption("Радио сурталчилгаа")
        render_editable_sheet("FM ads", current_user, FM_ADS_COLUMNS)
    
    # Cinema ads Tab
    with sheet_tabs[5]:
        st.markdown("### 🎬 Cinema ads")
        st.caption("Кино театрын сурталчилгаа")
        render_editable_sheet("Cinema ads", current_user, CINEMA_ADS_COLUMNS)
    
    # Shopping mall Tab
    with sheet_tabs[6]:
        st.markdown("### 🛒 Shopping mall; CU&GS25; Coffee")
        st.caption("Худалдааны төв, CU, GS25, Coffee shop")
        render_editable_sheet("Shopping mall", current_user, SHOPPING_MALL_COLUMNS)
    
    # Digital & Social Tab
    with sheet_tabs[7]:
        st.markdown("### 📱 Digital & Social")
        st.caption("Дижитал болон сошиал сурталчилгаа")
        render_editable_sheet("Digital & Social", current_user, DIGITAL_COLUMNS)


# =============================================================================
# MAIN PAGE
# =============================================================================

def main():
    """Main dashboard page."""
    
    # Check JWT authentication - Dashboard requires login
    jwt_user = get_current_user_from_token()
    if not jwt_user:
        st.title("📊 Marketing CPP Dashboard 2025")
        st.warning("⚠️ Dashboard-д нэвтрэх шаардлагатай")
        st.info("👈 Зүүн талын цэснээс **🏠 Home** хуудас руу очиж нэвтэрнэ үү.")
        
        # Quick login button redirect
        if st.button("🔐 Нэвтрэх хуудас руу очих"):
            st.switch_page("app.py")
        return
    
    # Get user from database for full object
    with get_session() as session:
        user = session.get(User, int(jwt_user['id']))
    
    if not user:
        st.error("Хэрэглэгч олдсонгүй. Дахин нэвтэрнэ үү.")
        return
    
    # Page header
    st.title("📊 Marketing CPP Dashboard 2025")
    st.sidebar.markdown(f"**Нэвтэрсэн:** {user.full_name or user.username}")
    
    # Create tabs for different views
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📈 CPP Analytics", "📋 Файлууд", "📑 Төсвийн тайлан", "📥 CPP Report", "📄 PDF Export"])
    
    with get_session() as session:
        with tab1:
            render_analytics_tab(session)
        
        with tab2:
            render_files_tab(session)
        
        with tab3:
            render_budget_report_tab(session)
        
        with tab4:
            render_cpp_report_tab(session)
        
        with tab5:
            render_export_tab(session)


# =============================================================================
# RUN PAGE
# =============================================================================

if __name__ == "__main__":
    main()
else:
    main()
