"""
CPP Report Generator
=====================

Generates the final CPP (Central Planning Platform) Report in Excel format.
The output file replicates the structure of the 2025_CPP.xlsx template.

This module:
1. Reads BudgetItem data from the database
2. Categorizes items by channel (TV, OOH, FM, etc.)
3. Generates a multi-sheet Excel file matching the CPP template format
4. Returns DataFrames for real-time preview in Dashboard

Author: CPP Development Team
"""

import os
from io import BytesIO
from typing import Dict, List, Any, Optional
from datetime import datetime
from decimal import Decimal

import pandas as pd
from sqlmodel import Session, select

from database.models import BudgetFile, BudgetItem, CppBudgetItem
from config import FileStatus, ChannelType


# =============================================================================
# COMPANY MAPPING
# =============================================================================

COMPANY_MAP = {
    'A': 'Юнител',
    'B': 'Юнивишн',
    'G': 'Green Future',
    'J': 'IVLBS',
    'T': 'MPSC'
}


# =============================================================================
# CPP SHEET COLUMN DEFINITIONS (Exact match to 2025_CPP.xlsx)
# =============================================================================

# Each sheet has exact column order matching 2025_CPP.xlsx

# TV Sheet - ТВ Сурталчилгааны Суваг
# Columns with formulas are calculated in Excel, UI shows them as read-only or auto-calculated
TV_ADS_COLUMNS = [
    '№',                        # Row number
    'Кампанит ажлын нэр',       # Dropdown - from budget files
    'Брэнд',                    # Auto-fill from campaign
    'Компани',                  # Auto-fill from campaign
    'Мэргэжилтэн',              # Auto-fill from logged-in user
    'Төсвийн код',              # Dropdown - from budget files
    'Үнийн дүн (₮)',            # Formula: =Нийт сурталчилгааны урт * 1 секундын үнэлгээ
    'Төсвийн эх үүсвэр',        # Dropdown: Кампанит ажил, Ивээн тэтгэлэг, Урт хугацааны суваг
    'Сурталчилгааны төрөл',     # Dropdown: Аман сурталчилгаа, Задгай цацалт, Нэвтрүүлэг
    'Сувгийн нэр',              # Dropdown - TV channels (MNB, TV9, NTV, etc.)
    'Эхлэх өдөр',               # Date
    'Дуусах өдөр',              # Date
    'Нийт өдөр',                # Formula: =Дуусах өдөр - Эхлэх өдөр
    'Сурталчилгааны цаг',       # Text: e.g., "19:00-23:00"
    'Сурталчилгааны урт',       # Number (seconds)
    'Нийт давтамж',             # Number
    'Нийт сурталчилгааны урт',  # Formula: =Сурталчилгааны урт * Нийт давтамж
    '1 секундын үнэлгээ (₮)',   # Number (₮)
    'Тайлбар',                  # Text/Comment
]

# Dropdown options for TV Sheet
TV_BUDGET_SOURCE_OPTIONS = ['Кампанит ажил', 'Ивээн тэтгэлэг', 'Урт хугацааны суваг']
TV_AD_TYPE_OPTIONS = ['Аман сурталчилгаа', 'Задгай цацалт', 'Нэвтрүүлэг']
TV_CHANNEL_OPTIONS = [
    'MNB', 'EDU', 'Монгол HD', 'Movie box', 'Asian box', 'Bloomberg',
    'Central TV', 'UBS', 'C1', 'MN25', 'TV5', 'TV9', 'SBN', 'NTV',
    'TV8', 'PSN', 'ETV', 'Малчин ТВ', 'Channel 11', 'Eagle', 'Эко',
    'Seven', 'Соён гэгээрүүлэгч', 'Like HD', 'Dream TV', 'Орон нутаг'
]

OOH_ADS_COLUMNS = [
    '№', 'Харилцагч', 'Компани', 'Brand', 'Кампанит ажлын нэр',
    'Мэргэжилтэн', 'Budget code', 'Нийт дүн', '1 өдрийн түрээсийн зардал',
    'Төрөл', 'Хэмжээ', 'Байршил', 'Start', 'End', 'Days',
    'Нэгж сек', '1 өдрийн давтамж', 'Нийт давтамж', 'Нийт сек', 'Тайлбар'
]

INDOOR_ADS_COLUMNS = [
    '№', 'Харилцагч', 'Компани', 'Brand', 'Кампанит ажлын нэр',
    'Мэргэжилтэн', 'Budget code', '1 өдрийн түрээсийн зардал', 'Нийт дүн',
    'Төрөл', 'Нийт самбар & Led тоо', 'Start', 'End', 'Days',
    'Нэгж сек', '1 өдрийн давтамж', 'Нийт давтамж', 'Нийт сек',
    'Хандалтын тоо', 'Тайлбар'
]

FM_ADS_COLUMNS = [
    '№', 'FM суваг', 'Компани', 'Brand', 'Кампанит ажлын нэр',
    'Мэргэжилтэн', 'Budget code', '1 өдрийн зардал', 'Нийт дүн',
    'Төрөл', 'Start', 'End', 'Days', 'Нэгж сек',
    '1 өдрийн давтамж', 'Нийт давтамж', 'Нийт сек', 'Тайлбар'
]

CINEMA_ADS_COLUMNS = [
    '№', 'Cinema', 'Компани', 'Brand', 'Кампанит ажлын нэр',
    'Мэргэжилтэн', 'Budget code', '1 өдрийн зардал', 'Нийт дүн',
    'Ads type', 'Байршил', 'Start', 'End', 'Days',
    'Нэгж сек', '1 өдрийн давтамж', 'Нийт давтамж', 'Нийт сек',
    'Нийт үзвэрийн тоо', 'Тайлбар'
]

SHOPPING_MALL_COLUMNS = [
    '№', 'Type', 'Brand', 'Кампанит ажлын нэр',
    'Мэргэжилтэн', 'Budget code', '1 өдрийн түрээсийн зардал', 'Нийт дүн',
    'Төрөл', 'Байршил', 'Start', 'End', 'Days',
    'Нэгж сек', '1 өдрийн давтамж', 'Нийт давтамж', 'Нийт сек',
    'Нийт үйлчлүүлэгчдийн тоо', 'Тайлбар'
]

DIGITAL_COLUMNS = [
    '№', 'Суваг', 'Компани', 'Brand', 'Кампанит ажлын нэр',
    'Мэргэжилтэн', 'Budget code', 'Нийт дүн', 'Төрөл',
    'Start', 'End', 'Days', 'Impressions', 'Clicks', 'Тайлбар'
]


# =============================================================================
# CHANNEL TYPE MAPPING
# =============================================================================

# Map ChannelType enum values to sheet names
# Available: TV, FM, OOH, DIGITAL, PRINT, EVENT, OTHER
CHANNEL_TO_SHEET = {
    ChannelType.TV: 'TV ads',
    ChannelType.OOH: 'OOH & DOOH ads',
    ChannelType.FM: 'FM ads',
    ChannelType.DIGITAL: 'Digital & Social',
    ChannelType.PRINT: 'Other',
    ChannelType.EVENT: 'Other',
    ChannelType.OTHER: 'Other',
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_company_from_code(budget_code: str) -> str:
    """Get company name from budget code prefix."""
    if not budget_code:
        return ''
    first_char = budget_code[0].upper()
    return COMPANY_MAP.get(first_char, '')


def safe_float(value) -> Optional[float]:
    """Convert value to float safely."""
    if value is None:
        return None
    try:
        if isinstance(value, Decimal):
            return float(value)
        return float(value)
    except (ValueError, TypeError):
        return None


def format_date(dt) -> str:
    """Format datetime to string."""
    if dt is None:
        return ''
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d')
    return str(dt)


def calculate_days(start_date, end_date) -> Optional[int]:
    """Calculate number of days between dates."""
    if start_date is None or end_date is None:
        return None
    try:
        if isinstance(start_date, datetime) and isinstance(end_date, datetime):
            return (end_date - start_date).days + 1
    except:
        pass
    return None


# =============================================================================
# MAIN DATA EXTRACTION FUNCTIONS
# =============================================================================

import glob
import openpyxl

def get_uploaded_excel_files(session: Session) -> List[Dict]:
    """
    Get list of uploaded Excel files with their paths.
    """
    files = session.exec(
        select(BudgetFile).where(
            BudgetFile.status.in_([
                FileStatus.PENDING_APPROVAL,
                FileStatus.APPROVED_FOR_PRINT,
                FileStatus.SIGNING,
                FileStatus.FINALIZED
            ])
        )
    ).all()
    
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    upload_dir = os.path.join(base_dir, 'assets', 'uploaded_files')
    
    result = []
    for f in files:
        # Find Excel file for this budget
        pattern = os.path.join(upload_dir, f"budget_{f.id}_*.xlsx")
        matches = glob.glob(pattern)
        
        if matches:
            excel_path = matches[-1]  # Latest file
            company = get_company_from_code(f.budget_code) if f.budget_code else ''
            
            result.append({
                'file_id': f.id,
                'budget_code': f.budget_code or '',
                'company': company,
                'brand': f.brand or company,
                'filename': f.filename,
                'excel_path': excel_path,
                'uploader_id': f.uploader_id,
                'total_amount': float(f.total_amount) if f.total_amount else 0,
                'planned_amount': float(f.planned_amount) if f.planned_amount else 0
            })
    
    return result


def parse_template_sheet(excel_path: str, file_info: Dict) -> List[Dict]:
    """
    Parse TEMPLATE sheet from uploaded Excel file.
    Returns list of row dictionaries.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Find TEMPLATE sheet
        template_sheet = None
        for sn in wb.sheetnames:
            if sn.upper() == 'TEMPLATE':
                template_sheet = sn
                break
        
        if not template_sheet:
            return []
        
        ws = wb[template_sheet]
        
        # Find header row (look for "хийгдэх ажил" or similar)
        header_row = None
        headers = {}
        
        for row_idx in range(1, min(30, ws.max_row + 1)):
            row_values = [str(ws.cell(row_idx, col).value or '').strip().lower() for col in range(1, min(30, ws.max_column + 1))]
            
            # Check for key header keywords
            if any(kw in ' '.join(row_values) for kw in ['хийгдэх ажил', 'төрөл', 'нийт төсөв', 'channel']):
                header_row = row_idx
                for col_idx, val in enumerate(row_values, 1):
                    if val:
                        headers[col_idx] = val
                break
        
        if not header_row:
            return []
        
        # Parse data rows
        rows = []
        for row_idx in range(header_row + 1, min(ws.max_row + 1, 200)):
            row_data = {}
            has_data = False
            
            for col_idx, header_name in headers.items():
                cell_val = ws.cell(row_idx, col_idx).value
                if cell_val is not None and str(cell_val).strip():
                    has_data = True
                row_data[header_name] = cell_val
            
            if has_data:
                # Add file metadata
                row_data['_file_id'] = file_info['file_id']
                row_data['_company'] = file_info['company']
                row_data['_brand'] = file_info['brand']
                row_data['_budget_code'] = file_info['budget_code']
                row_data['_uploader_id'] = file_info['uploader_id']
                rows.append(row_data)
        
        wb.close()
        return rows
        
    except Exception as e:
        print(f"Error parsing {excel_path}: {e}")
        return []


def get_all_template_data(session: Session) -> List[Dict]:
    """
    Get all data from TEMPLATE sheets of uploaded Excel files.
    """
    excel_files = get_uploaded_excel_files(session)
    
    all_data = []
    for file_info in excel_files:
        rows = parse_template_sheet(file_info['excel_path'], file_info)
        all_data.extend(rows)
    
    return all_data


def categorize_by_channel(row: Dict) -> str:
    """
    Determine which CPP sheet this row belongs to based on content.
    """
    # Check all values for channel keywords
    row_text = ' '.join(str(v).upper() for v in row.values() if v and not str(v).startswith('_'))
    
    # TV keywords
    if any(kw in row_text for kw in ['ТВ СУВАГ', 'TV CHANNEL', 'ASIAN BOX', 'MOVIE BOX', 'EDU', 'NTV', 'CENTRAL TV', 'ТВ СУРТАЛЧИЛГАА']):
        return 'TV ads'
    
    # FM keywords
    if any(kw in row_text for kw in ['FM', 'РАДИО', 'RADIO', 'MGL FM', 'MGL 88.3']):
        return 'FM ads'
    
    # Cinema keywords
    if any(kw in row_text for kw in ['КИНО', 'CINEMA', 'PRIMECINEPLEX', 'ҮЗВЭР']):
        return 'Cinema ads'
    
    # Indoor keywords
    if any(kw in row_text for kw in ['ДОТООД', 'INDOOR', 'ЛИФТ', 'LED ДОТОР', 'ХҮРЭЭ ДИЗАЙН', 'ХАППИ']):
        return 'Indoor ads'
    
    # Shopping mall keywords
    if any(kw in row_text for kw in ['CU', 'GS25', 'SHOPPING', 'MALL', 'EMART', 'COFFEE', 'ТҮРДЭГ ТЭРЭГ']):
        return 'Shopping mall'
    
    # OOH keywords
    if any(kw in row_text for kw in ['ГАДНАХ', 'OOH', 'DOOH', 'САМБАР', 'BILLBOARD', 'LED ГАДНА', 'СИПИМЕДИА']):
        return 'OOH & DOOH ads'
    
    # Digital keywords
    if any(kw in row_text for kw in ['ДИЖИТАЛ', 'DIGITAL', 'СОШИАЛ', 'SOCIAL', 'FACEBOOK', 'INSTAGRAM', 'YOUTUBE', 'INFLUENCER']):
        return 'Digital & Social'
    
    # Default to OOH
    return 'OOH & DOOH ads'


def get_file_info_map(session: Session) -> Dict[int, Dict]:
    """
    Get file info map for looking up company, specialist etc.
    """
    files = session.exec(select(BudgetFile)).all()
    
    result = {}
    for f in files:
        company = get_company_from_code(f.budget_code) if f.budget_code else ''
        result[f.id] = {
            'budget_code': f.budget_code or '',
            'company': company,
            'brand': f.brand or company,
            'filename': f.filename,
            'uploader_id': f.uploader_id
        }
    
    return result


# =============================================================================
# DATAFRAME GENERATION FUNCTIONS
# =============================================================================

def get_cpp_report_dataframes(session: Session) -> Dict[str, pd.DataFrame]:
    """
    Generate all CPP report DataFrames from uploaded Excel files.
    Reads TEMPLATE sheets directly and categorizes by channel.
    
    Returns:
        Dictionary with sheet names as keys and DataFrames as values.
        Each DataFrame has exact column order matching 2025_CPP.xlsx.
    """
    # Get all data from TEMPLATE sheets
    all_rows = get_all_template_data(session)
    
    # Initialize data containers for each sheet
    sheet_data = {
        'TV ads': [],
        'OOH & DOOH ads': [],
        'Indoor ads': [],
        'FM ads': [],
        'Cinema ads': [],
        'Shopping mall': [],
        'Digital & Social': []
    }
    
    # Process each row and categorize
    for row in all_rows:
        sheet_name = categorize_by_channel(row)
        
        # Extract common fields from row
        company = row.get('_company', '')
        brand = row.get('_brand', company)
        budget_code = row.get('_budget_code', '')
        file_id = row.get('_file_id', 0)
        uploader_id = row.get('_uploader_id', 0)
        
        # Try to extract values from various possible column names
        campaign_name = (
            row.get('хийгдэх ажил') or 
            row.get('кампанит ажлын нэр') or 
            row.get('campaign') or 
            ''
        )
        vendor = (
            row.get('харилцагч') or 
            row.get('vendor') or
            row.get('суваг') or
            ''
        )
        amount = safe_float(
            row.get('нийт төсөв') or 
            row.get('total budget') or 
            row.get('бодит') or
            row.get('нийт дүн') or
            0
        )
        description = (
            row.get('тайлбар') or 
            row.get('description') or 
            ''
        )
        activity_type = (
            row.get('төрөл') or 
            row.get('type') or
            ''
        )
        
        # Create row based on sheet type
        if sheet_name == 'TV ads':
            sheet_data['TV ads'].append({
                '№': len(sheet_data['TV ads']) + 1,
                'Кампанит ажлын нэр': campaign_name,
                'Компани': company,
                'Мэргэжилтэн': '',
                'Төсвийн код': budget_code,
                'TV ads ACTUAL': amount,
                'Сурталчилгааны төрөл': activity_type,
                'TV нэр': vendor,
                'Scope': '',
                'Start date': '',
                'End date': '',
                'Days': '',
                'Unit seconds': '',
                'Total frequency': '',
                'Total seconds': '',
                'Цаг': '',
                'Description': description,
                '_file_id': file_id,
                '_uploader_id': uploader_id
            })
            
        elif sheet_name == 'OOH & DOOH ads':
            sheet_data['OOH & DOOH ads'].append({
                '№': len(sheet_data['OOH & DOOH ads']) + 1,
                'Харилцагч': vendor,
                'Компани': company,
                'Brand': brand,
                'Кампанит ажлын нэр': campaign_name,
                'Мэргэжилтэн': '',
                'Budget code': budget_code,
                'Нийт дүн': amount,
                '1 өдрийн түрээсийн зардал': '',
                'Төрөл': activity_type,
                'Хэмжээ': '',
                'Байршил': description,
                'Start': '',
                'End': '',
                'Days': '',
                'Нэгж сек': '',
                '1 өдрийн давтамж': '',
                'Нийт давтамж': '',
                'Нийт сек': '',
                'Тайлбар': '',
                '_file_id': file_id,
                '_uploader_id': uploader_id
            })
            
        elif sheet_name == 'Indoor ads':
            sheet_data['Indoor ads'].append({
                '№': len(sheet_data['Indoor ads']) + 1,
                'Харилцагч': vendor,
                'Компани': company,
                'Brand': brand,
                'Кампанит ажлын нэр': campaign_name,
                'Мэргэжилтэн': '',
                'Budget code': budget_code,
                '1 өдрийн түрээсийн зардал': '',
                'Нийт дүн': amount,
                'Төрөл': activity_type,
                'Нийт самбар & Led тоо': '',
                'Start': '',
                'End': '',
                'Days': '',
                'Нэгж сек': '',
                '1 өдрийн давтамж': '',
                'Нийт давтамж': '',
                'Нийт сек': '',
                'Хандалтын тоо': '',
                'Тайлбар': description,
                '_file_id': file_id,
                '_uploader_id': uploader_id
            })
            
        elif sheet_name == 'FM ads':
            sheet_data['FM ads'].append({
                '№': len(sheet_data['FM ads']) + 1,
                'FM суваг': vendor,
                'Компани': company,
                'Brand': brand,
                'Кампанит ажлын нэр': campaign_name,
                'Мэргэжилтэн': '',
                'Budget code': budget_code,
                '1 өдрийн зардал': '',
                'Нийт дүн': amount,
                'Төрөл': activity_type,
                'Start': '',
                'End': '',
                'Days': '',
                'Нэгж сек': '',
                '1 өдрийн давтамж': '',
                'Нийт давтамж': '',
                'Нийт сек': '',
                'Тайлбар': description,
                '_file_id': file_id,
                '_uploader_id': uploader_id
            })
            
        elif sheet_name == 'Cinema ads':
            sheet_data['Cinema ads'].append({
                '№': len(sheet_data['Cinema ads']) + 1,
                'Cinema': vendor,
                'Компани': company,
                'Brand': brand,
                'Кампанит ажлын нэр': campaign_name,
                'Мэргэжилтэн': '',
                'Budget code': budget_code,
                '1 өдрийн зардал': '',
                'Нийт дүн': amount,
                'Ads type': activity_type,
                'Байршил': description,
                'Start': '',
                'End': '',
                'Days': '',
                'Нэгж сек': '',
                '1 өдрийн давтамж': '',
                'Нийт давтамж': '',
                'Нийт сек': '',
                'Нийт үзвэрийн тоо': '',
                'Тайлбар': '',
                '_file_id': file_id,
                '_uploader_id': uploader_id
            })
            
        elif sheet_name == 'Shopping mall':
            sheet_data['Shopping mall'].append({
                '№': len(sheet_data['Shopping mall']) + 1,
                'Type': vendor,
                'Brand': brand,
                'Кампанит ажлын нэр': campaign_name,
                'Мэргэжилтэн': '',
                'Budget code': budget_code,
                '1 өдрийн түрээсийн зардал': '',
                'Нийт дүн': amount,
                'Төрөл': activity_type,
                'Байршил': description,
                'Start': '',
                'End': '',
                'Days': '',
                'Нэгж сек': '',
                '1 өдрийн давтамж': '',
                'Нийт давтамж': '',
                'Нийт сек': '',
                'Нийт үйлчлүүлэгчдийн тоо': '',
                'Тайлбар': '',
                '_file_id': file_id,
                '_uploader_id': uploader_id
            })
            
        else:  # Digital & Social
            sheet_data['Digital & Social'].append({
                '№': len(sheet_data['Digital & Social']) + 1,
                'Суваг': vendor,
                'Компани': company,
                'Brand': brand,
                'Кампанит ажлын нэр': campaign_name,
                'Мэргэжилтэн': '',
                'Budget code': budget_code,
                'Нийт дүн': amount,
                'Төрөл': activity_type,
                'Start': '',
                'End': '',
                'Days': '',
                'Impressions': '',
                'Clicks': '',
                'Тайлбар': description,
                '_file_id': file_id,
                '_uploader_id': uploader_id
            })
    
    # Create DataFrames with metadata columns (for editing)
    result = {}
    
    column_maps = {
        'TV ads': TV_ADS_COLUMNS + ['_file_id', '_uploader_id'],
        'OOH & DOOH ads': OOH_ADS_COLUMNS + ['_file_id', '_uploader_id'],
        'Indoor ads': INDOOR_ADS_COLUMNS + ['_file_id', '_uploader_id'],
        'FM ads': FM_ADS_COLUMNS + ['_file_id', '_uploader_id'],
        'Cinema ads': CINEMA_ADS_COLUMNS + ['_file_id', '_uploader_id'],
        'Shopping mall': SHOPPING_MALL_COLUMNS + ['_file_id', '_uploader_id'],
        'Digital & Social': DIGITAL_COLUMNS + ['_file_id', '_uploader_id']
    }
    
    for sheet_name, data in sheet_data.items():
        if data:
            df = pd.DataFrame(data)
            # Reorder columns
            cols = column_maps.get(sheet_name, list(df.columns))
            existing_cols = [c for c in cols if c in df.columns]
            result[sheet_name] = df[existing_cols]
        else:
            # Empty dataframe with correct columns
            cols = column_maps.get(sheet_name, [])
            result[sheet_name] = pd.DataFrame(columns=cols)
    
    return result


def create_general_sheet(session: Session) -> pd.DataFrame:
    """
    Create the General summary sheet.
    """
    files = session.exec(
        select(BudgetFile).where(
            BudgetFile.status.in_([
                FileStatus.PENDING_APPROVAL,
                FileStatus.APPROVED_FOR_PRINT,
                FileStatus.SIGNING,
                FileStatus.FINALIZED
            ])
        )
    ).all()
    
    # Calculate totals by company
    company_totals = {}
    for f in files:
        company = get_company_from_code(f.budget_code) if f.budget_code else 'Other'
        if company not in company_totals:
            company_totals[company] = {'planned': 0, 'actual': 0}
        
        if f.planned_amount:
            company_totals[company]['planned'] += float(f.planned_amount)
        if f.total_amount:
            company_totals[company]['actual'] += float(f.total_amount)
    
    # Create summary data
    summary_data = []
    for company, totals in company_totals.items():
        summary_data.append({
            'Компани': company,
            'Нийт төсөв': totals['planned'],
            'Бодит зарцуулалт': totals['actual'],
            'Зөрүү': totals['planned'] - totals['actual']
        })
    
    return pd.DataFrame(summary_data)


# =============================================================================
# EXCEL EXPORT FUNCTIONS
# =============================================================================

def export_cpp_report(session: Session) -> bytes:
    """
    Generate CPP Report Excel file and return as bytes.
    """
    dataframes = get_cpp_report_dataframes(session)
    general_df = create_general_sheet(session)
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        
        # Define formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        
        number_format = workbook.add_format({
            'num_format': '#,##0',
            'border': 1
        })
        
        date_format = workbook.add_format({
            'num_format': 'yyyy-mm-dd',
            'border': 1
        })
        
        text_format = workbook.add_format({
            'border': 1,
            'text_wrap': True
        })
        
        # Write General sheet first
        general_df.to_excel(writer, sheet_name='General', index=False, startrow=1)
        ws_general = writer.sheets['General']
        
        # Write title
        ws_general.write(0, 0, 'CPP Report Summary', workbook.add_format({
            'bold': True, 'font_size': 14
        }))
        
        # Format General sheet headers
        for col_num, col_name in enumerate(general_df.columns):
            ws_general.write(1, col_num, col_name, header_format)
        
        # Set column widths for General
        ws_general.set_column(0, 0, 15)  # Company
        ws_general.set_column(1, 3, 18)  # Amounts
        
        # Write each channel sheet
        sheet_order = ['TV ads', 'OOH & DOOH ads', 'Indoor ads', 'FM ads', 'Cinema ads', 'Shopping mall', 'Digital & Social']
        
        for sheet_name in sheet_order:
            if sheet_name in dataframes:
                df = dataframes[sheet_name].copy()
                
                # Remove metadata columns for export
                export_cols = [c for c in df.columns if not c.startswith('_')]
                df = df[export_cols]
                
                # Determine start row (matching 2025_CPP.xlsx)
                if sheet_name == 'TV ads':
                    start_row = 8  # Header at row 9 (0-indexed = 8)
                elif sheet_name == 'Cinema ads':
                    start_row = 6  # Header at row 7
                elif sheet_name == 'Shopping mall':
                    start_row = 5  # Header at row 6
                else:
                    start_row = 2  # Header at row 3
                
                # Write DataFrame
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
                
                ws = writer.sheets[sheet_name]
                
                # Write headers with format
                for col_num, col_name in enumerate(df.columns):
                    ws.write(start_row, col_num, col_name, header_format)
                
                # Set column widths
                for col_num, col_name in enumerate(df.columns):
                    if '№' in col_name:
                        ws.set_column(col_num, col_num, 5)
                    elif 'дүн' in col_name.lower() or 'actual' in col_name.lower() or 'зардал' in col_name.lower():
                        ws.set_column(col_num, col_num, 18)
                    elif 'нэр' in col_name.lower() or 'description' in col_name.lower() or 'тайлбар' in col_name.lower():
                        ws.set_column(col_num, col_num, 30)
                    elif 'date' in col_name.lower() or col_name in ['Start', 'End']:
                        ws.set_column(col_num, col_num, 12)
                    else:
                        ws.set_column(col_num, col_num, 15)
                
                # Freeze panes
                ws.freeze_panes(start_row + 1, 0)
    
    output.seek(0)
    return output.getvalue()


# =============================================================================
# LEGACY FUNCTION (for backward compatibility)
# =============================================================================

def generate_cpp_report(session: Session) -> bytes:
    """
    Legacy function - calls export_cpp_report.
    """
    return export_cpp_report(session)


# =============================================================================
# CPP ITEMS EXPORT (Manual entry data from CppBudgetItem)
# =============================================================================

def export_cpp_items_to_excel(session: Session) -> bytes:
    """
    Export CppBudgetItem data to Excel.
    Each category becomes a separate sheet.
    
    Returns:
        bytes: Excel file content
    """
    output = BytesIO()
    
    # Define sheets and their columns
    sheets_config = {
        'TV ads': TV_ADS_COLUMNS,
        'OOH & DOOH ads': OOH_ADS_COLUMNS,
        'Indoor ads': INDOOR_ADS_COLUMNS,
        'FM ads': FM_ADS_COLUMNS,
        'Cinema ads': CINEMA_ADS_COLUMNS,
        'Shopping mall': SHOPPING_MALL_COLUMNS,
        'Digital & Social': DIGITAL_COLUMNS,
    }
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        
        # Create formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center'
        })
        
        cell_format = workbook.add_format({
            'border': 1,
            'valign': 'vcenter'
        })
        
        number_format = workbook.add_format({
            'border': 1,
            'valign': 'vcenter',
            'num_format': '#,##0'
        })
        
        for sheet_name, columns in sheets_config.items():
            # Get items for this category
            stmt = select(CppBudgetItem).where(
                CppBudgetItem.category_name == sheet_name
            ).order_by(CppBudgetItem.row_number)
            items = session.exec(stmt).all()
            
            # Build dataframe rows
            rows = []
            for idx, item in enumerate(items, start=1):
                row = {'№': idx}
                custom = item.custom_fields or {}
                
                # Parse custom_fields if it's a string
                if isinstance(custom, str):
                    try:
                        import json
                        custom = json.loads(custom)
                    except:
                        custom = {}
                
                if not isinstance(custom, dict):
                    custom = {}
                
                # Fill each column from custom_fields
                for col in columns:
                    if col == '№':
                        continue
                    row[col] = custom.get(col, '')
                
                rows.append(row)
            
            # Create dataframe
            if rows:
                df = pd.DataFrame(rows, columns=columns)
            else:
                # Empty dataframe with columns
                df = pd.DataFrame(columns=columns)
            
            # Write to Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            
            ws = writer.sheets[sheet_name]
            
            # Write header
            for col_num, col_name in enumerate(columns):
                ws.write(0, col_num, col_name, header_format)
            
            # Set column widths
            for col_num, col_name in enumerate(columns):
                if '№' in col_name:
                    ws.set_column(col_num, col_num, 5)
                elif 'дүн' in col_name.lower() or 'зардал' in col_name.lower() or 'actual' in col_name.lower():
                    ws.set_column(col_num, col_num, 18)
                elif 'нэр' in col_name.lower() or 'description' in col_name.lower() or 'тайлбар' in col_name.lower():
                    ws.set_column(col_num, col_num, 30)
                elif 'date' in col_name.lower() or col_name in ['Start', 'End']:
                    ws.set_column(col_num, col_num, 12)
                else:
                    ws.set_column(col_num, col_num, 15)
            
            # Freeze panes
            ws.freeze_panes(1, 0)
    
    output.seek(0)
    return output.getvalue()
